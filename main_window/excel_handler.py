# main_window/excel_handler.py
# Excel file handling module for NSIS application

import os
import datetime
import logging
from typing import List, Dict, Any, Optional, Tuple
from PyQt6 import QtWidgets

# Import configuration
from config import (
    COL_RICERCA, COL_STATO, COL_PROTOCOLLO, COL_PROVVEDIMENTO,
    COL_DATA_PROVV, COL_CODICE_RIS, COL_NOTE
)

class ExcelHandler:
    """Handles Excel file operations for the NSIS application."""
    
    def __init__(self):
        self._logger = logging.getLogger(__name__)
        self._current_file_path: Optional[str] = None
        self._codes: List[str] = []
        
    def load_excel_file(self, file_path: str) -> Tuple[bool, List[str], str]:
        """
        Load Excel file and extract codes from the search column.
        
        Returns:
            Tuple[bool, List[str], str]: (success, codes_list, error_message)
        """
        self._current_file_path = file_path
        
        try:
            import openpyxl
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError:
            error_msg = "Libreria 'openpyxl' non trovata. Installala con 'pip install openpyxl'"
            self._logger.error(error_msg)
            return False, [], error_msg
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
            
            if not workbook.sheetnames:
                error_msg = "File Excel non contiene fogli di lavoro."
                self._logger.error(error_msg)
                return False, [], error_msg
            
            sheet = workbook.active
            self._logger.info(f"Caricato foglio: {sheet.title}")
            
            # Find search column
            search_col_idx = self._find_search_column(sheet)
            if search_col_idx is None:
                error_msg = f"Colonna di ricerca '{COL_RICERCA}' non trovata nel file."
                self._logger.error(error_msg)
                return False, [], error_msg
            
            # Extract codes
            codes = self._extract_codes_from_column(sheet, search_col_idx)
            
            workbook.close()
            
            self._codes = codes
            self._logger.info(f"Estratti {len(codes)} codici dal file Excel")
            
            return True, codes, ""
            
        except (PermissionError, IOError) as e:
            error_msg = f"Errore di accesso al file: {e}"
            self._logger.error(error_msg)
            return False, [], error_msg
        except InvalidFileException as e:
            error_msg = f"File Excel non valido: {e}"
            self._logger.error(error_msg)
            return False, [], error_msg
        except Exception as e:
            error_msg = f"Errore imprevisto durante il caricamento: {e}"
            self._logger.error(error_msg, exc_info=True)
            return False, [], error_msg
    
    def _find_search_column(self, sheet) -> Optional[int]:
        """Find the search column index in the Excel sheet."""
        header_row = sheet[1]  # First row contains headers
        
        for cell in header_row:
            if cell.value and str(cell.value).strip().lower() == COL_RICERCA.lower():
                return cell.column
        
        return None
    
    def _extract_codes_from_column(self, sheet, column_idx: int) -> List[str]:
        """Extract codes from the specified column."""
        codes = []
        
        for row in sheet.iter_rows(min_row=2, min_col=column_idx, max_col=column_idx):
            cell = row[0]
            if cell.value:
                code = str(cell.value).strip()
                if code:  # Skip empty strings
                    codes.append(code)
        
        return codes
    
    def save_results_to_excel(self, results_list: List[Dict[str, Any]], 
                            original_file_path: str) -> Tuple[bool, str, str]:
        """
        Save results to Excel file.
        
        Returns:
            Tuple[bool, str, str]: (success, output_file_path, error_message)
        """
        try:
            import openpyxl
            from openpyxl.utils.exceptions import InvalidFileException
            from openpyxl.styles import Font, Alignment, numbers
            from openpyxl.utils import get_column_letter
        except ImportError:
            error_msg = "Libreria 'openpyxl' non trovata. Impossibile salvare file Excel."
            self._logger.error(error_msg)
            return False, "", error_msg
        
        self._logger.info(f"Tentativo salvataggio {len(results_list)} risultati su {os.path.basename(original_file_path)}")
        
        workbook = None
        output_file_path = original_file_path
        is_read_only_or_corrupted = False
        
        try:
            # 1. Try to load workbook for writing
            try:
                workbook = openpyxl.load_workbook(filename=original_file_path)
                self._logger.debug(f"Workbook '{os.path.basename(original_file_path)}' caricato per scrittura.")
                is_read_only_or_corrupted = False
                output_file_path = original_file_path
            except (PermissionError, IOError) as pe:
                self._logger.warning(f"Errore permesso/IO su '{os.path.basename(original_file_path)}': {pe}. Tento salvataggio su copia.")
                is_read_only_or_corrupted = True
                output_file_path = self._create_backup_file(original_file_path)
                if not output_file_path:
                    return False, "", "Impossibile creare copia del file"
                workbook = openpyxl.load_workbook(filename=output_file_path)
            except (InvalidFileException, KeyError, Exception) as e_load:
                self._logger.error(f"Errore caricamento workbook '{os.path.basename(original_file_path)}': {e_load}")
                return False, "", f"Impossibile caricare file Excel: {e_load}"
            
            # Get active sheet
            if not workbook or not workbook.sheetnames:
                return False, "", "Workbook non valido o senza fogli."
            
            sheet = workbook.active
            
            # Process results and write to sheet
            success = self._write_results_to_sheet(sheet, results_list)
            
            if success:
                # Save workbook
                workbook.save(output_file_path)
                self._logger.info(f"Risultati salvati con successo in: {output_file_path}")
                return True, output_file_path, ""
            else:
                return False, "", "Errore durante la scrittura dei risultati"
                
        except Exception as e:
            error_msg = f"Errore durante il salvataggio: {e}"
            self._logger.error(error_msg, exc_info=True)
            return False, "", error_msg
        finally:
            if workbook:
                try:
                    workbook.close()
                except:
                    pass
    
    def _create_backup_file(self, original_file_path: str) -> Optional[str]:
        """Create a backup file when original is read-only."""
        try:
            import openpyxl
            
            # Load original in read-only mode
            workbook_ro = openpyxl.load_workbook(filename=original_file_path, read_only=True, data_only=True)
            if not workbook_ro.sheetnames:
                return None
            
            original_sheet = workbook_ro.active
            
            # Create new workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = original_sheet.title
            
            # Copy header and data
            for row_idx in range(1, original_sheet.max_row + 1):
                row_values = [cell.value for cell in original_sheet[row_idx]]
                sheet.append(row_values)
            
            workbook_ro.close()
            
            # Create backup filename
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(original_file_path)
            backup_path = f"{base}_output_{timestamp}{ext}"
            
            # Save backup
            workbook.save(backup_path)
            workbook.close()
            
            self._logger.info(f"Backup creato: {backup_path}")
            return backup_path
            
        except Exception as e:
            self._logger.error(f"Errore creazione backup: {e}")
            return None
    
    def _write_results_to_sheet(self, sheet, results_list: List[Dict[str, Any]]) -> bool:
        """Write results to the Excel sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            
            # Column mapping
            col_name_map = {
                'Input Code': COL_RICERCA,
                'Stato': COL_STATO,
                'Protocollo uscita': COL_PROTOCOLLO,
                'Provvedimento': COL_PROVVEDIMENTO,
                'Data Provvedimento': COL_DATA_PROVV,
                'Codice richiesta (risultato)': COL_CODICE_RIS,
                'Note Usmaf': COL_NOTE
            }
            
            output_config_keys = [
                'Stato', 'Protocollo uscita', 'Provvedimento',
                'Data Provvedimento', 'Codice richiesta (risultato)', 'Note Usmaf'
            ]
            
            # Get existing headers
            header_row = sheet[1]
            existing_headers = {str(cell.value).strip().lower(): cell.column for cell in header_row if cell.value}
            
            # Find search column
            ricerca_col_name_lower = COL_RICERCA.lower()
            if ricerca_col_name_lower not in existing_headers:
                self._logger.error(f"Colonna di ricerca '{COL_RICERCA}' non trovata")
                return False
            
            ricerca_col_idx = existing_headers[ricerca_col_name_lower]
            
            # Determine next available column
            next_available_col_idx = sheet.max_column + 1
            
            # Map column indices
            final_col_indices = {}
            header_font = Font(bold=True)
            
            # Find or create output columns
            for config_key in output_config_keys:
                excel_header_name = col_name_map[config_key]
                excel_header_name_lower = excel_header_name.lower()
                
                if excel_header_name_lower in existing_headers:
                    # Column exists, use its index
                    col_idx = existing_headers[excel_header_name_lower]
                    final_col_indices[config_key] = col_idx
                else:
                    # Column missing, add it
                    col_idx = next_available_col_idx
                    final_col_indices[config_key] = col_idx
                    
                    # Add header
                    header_cell = sheet.cell(row=1, column=col_idx)
                    header_cell.value = excel_header_name
                    header_cell.font = header_font
                    
                    next_available_col_idx += 1
            
            # Write results
            for result in results_list:
                # Find row for this code
                code = result.get('Input Code', '')
                self._logger.debug(f"Processando risultato per codice: '{code}'")
                row_idx = self._find_row_for_code(sheet, ricerca_col_idx, code)
                
                if row_idx is None:
                    self._logger.warning(f"Riga non trovata per codice: {code}")
                    continue
                
                # Write result data
                for config_key in output_config_keys:
                    if config_key in final_col_indices:
                        col_idx = final_col_indices[config_key]
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.value = result.get(config_key, '')
            
            return True
            
        except Exception as e:
            self._logger.error(f"Errore scrittura risultati: {e}", exc_info=True)
            return False
    
    def _find_row_for_code(self, sheet, search_col_idx: int, code: str) -> Optional[int]:
        """Find the row index for a specific code."""
        self._logger.debug(f"Ricerca codice '{code}' nella colonna {search_col_idx}")
        for row_idx in range(2, sheet.max_row + 1):  # Skip header row
            cell = sheet.cell(row=row_idx, column=search_col_idx)
            cell_value = cell.value
            if cell_value:
                cell_value_str = str(cell_value).strip()
                self._logger.debug(f"Riga {row_idx}: confronto '{code}' con '{cell_value_str}'")
                if cell_value_str == code.strip():
                    self._logger.debug(f"Codice trovato nella riga {row_idx}")
                    return row_idx
        self._logger.warning(f"Codice '{code}' non trovato in nessuna riga")
        return None
    
    @property
    def current_file_path(self) -> Optional[str]:
        """Get current file path."""
        return self._current_file_path
    
    @property
    def codes(self) -> List[str]:
        """Get loaded codes."""
        return self._codes.copy()
    
    def reset(self):
        """Reset handler state."""
        self._current_file_path = None
        self._codes = [] 