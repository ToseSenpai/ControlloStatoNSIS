import os
import sys
import time
import pandas as pd
import math

from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QVBoxLayout,
    QFileDialog, QLabel, QProgressBar, QHBoxLayout, QTextEdit, QGridLayout, QFrame,
    QGraphicsColorizeEffect, QGraphicsOpacityEffect, QSizePolicy, QSpacerItem,
    QScrollBar
)
from PyQt5.QtCore import (
    Qt, QEventLoop, QObject, pyqtSlot, QUrl,
    QPropertyAnimation, QTimer, QRectF
)
from PyQt5.QtGui import (
    QPainter, QColor, QBrush, QPen, QPaintEvent,
    QLinearGradient, QGradient
)

try:
    from PyQt5.QtWebEngineWidgets import QWebEngineView, QWebEnginePage
except ImportError:
    raise ImportError(
        "Il modulo PyQtWebEngine non è installato.\n"
        "Installa eseguendo: pip install PyQtWebEngine"
    )
from PyQt5.QtWebChannel import QWebChannel

from openpyxl import load_workbook
from openpyxl.styles import numbers

STATO_SELECTOR = "#risultatiConsultazionePratica tbody tr td:nth-child(3)"
ALL_CELLS_JS = '''(function() {
    var row = document.querySelector('#risultatiConsultazionePratica tbody tr');
    if(!row) return null;
    var tds = row.querySelectorAll('td');
    var texts = [];
    tds.forEach(function(td){ texts.push(td.innerText.trim()); });
    return texts;
})();'''

class WebEnginePage(QWebEnginePage):
    def acceptNavigationRequest(self, url, nav_type, is_main_frame):
        return True

class JSBridge(QObject):
    def __init__(self):
        super().__init__()
        self.result = None
        self._loop = None

    @pyqtSlot(object)
    def receive(self, data):
        self.result = data
        if self._loop:
            self._loop.quit()

    def evaluate(self, page, script):
        self.result = None
        loop = QEventLoop()
        self._loop = loop
        page.runJavaScript(script, self.receive)
        loop.exec_()
        self._loop = None
        return self.result

# --- Classe Personalizzata per la Progress Bar con Animazione (Metodo Aggiornato) ---
class CustomProgressBar(QProgressBar):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Inizializza lo stato dell'animazione
        self._animation_phase = 0

        # Configura un timer per l'animazione (~60 FPS)
        self._animation_timer = QTimer(self)
        self._animation_timer.timeout.connect(self.animate)
        # *** RI-ABILITATO: Avvia il timer ***
        self._animation_timer.start(16)


        self.setStyleSheet("""
            QProgressBar {
                border: 1px solid #bdc3c7; /* Bordo del container */
                border-radius: 5px;
                background-color: #ecf0f1; /* Sfondo della parte vuota */
                text-align: center;
                color: #333333; /* Colore testo (nascosto) */
            }
            QProgressBar::chunk {
                 background-color: transparent; /* Disegniamo noi la chunk */
            }
        """)
        self.setTextVisible(False)

    def animate(self):
        # Aggiorna la fase dell'animazione (incrementa e ripete)
        # La velocità e la lunghezza dell'animazione dipendono da questo incremento e dal modulo (%)
        self._animation_phase = (self._animation_phase + 2) % 200 # Incremento e lunghezza ciclo animazione
        self.update() # Richiede un ridisegno

    # Override del metodo di disegno - Disegno con Gradiente Animato (Metodo Spostamento Painter)
    def paintEvent(self, event: QPaintEvent):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        # Rimosso painter.translate(self.rect().topLeft())

        rect = self.rect()
        borderRadius = 5

        # --- Disegno del background arrotondato con bordo ---
        painter.setBrush(QBrush(QColor("#ecf0f1")))
        painter.setPen(QPen(QColor("#bdc3c7"), 1))
        # Disegna il rettangolo esterno con bordi arrotondati
        painter.drawRoundedRect(rect, borderRadius, borderRadius)


        # --- Disegno della parte riempita (chunk) ---
        if self.maximum() > self.minimum():
            # Calcola la larghezza della parte riempita basata sul valore corrente
            # Riduciamo la larghezza totale di 2px per tenere conto del bordo
            barWidth = rect.width() - 2
            chunkWidth = int(((self.value() - self.minimum()) / (self.maximum() - self.minimum())) * barWidth)

            if chunkWidth > 0:
                # Definisci il rettangolo che rappresenta la parte riempita (interno al bordo)
                chunkFillRect = QRectF(rect.left() + 1, rect.top() + 1, chunkWidth, rect.height() - 2)

                # --- Gradiente Animato (Spostamento del Painter) ---
                # Definiamo un gradiente lineare che va da x=0 a x=50 (un pattern di 50px)
                # Questo gradiente verrà poi "spostato" applicando una traslazione al painter
                gradient = QLinearGradient(0, 0, 50, 0) # Gradiente definito su 50px di lunghezza

                # Definiamo gli stop del gradiente (es. blu-scuro -> blu-chiaro -> blu-scuro)
                gradient.setColorAt(0.0, QColor("#2980b9")) # Blu scuro all'inizio del pattern
                gradient.setColorAt(0.5, QColor("#3498db")) # Blu chiaro a metà
                gradient.setColorAt(1.0, QColor("#2980b9")) # Blu scuro alla fine del pattern

                # Imposta la modalità di ripetizione per coprire l'intera area del chunk
                gradient.setSpread(QGradient.RepeatSpread)

                # Calcola lo spostamento basato sulla fase di animazione
                # Mappiamo la fase (0-199) su uno spostamento che si ripete sulla lunghezza del gradiente pattern (50px)
                offset = (self._animation_phase / 199.0) * 50.0 # Spostamento massimo = lunghezza pattern

                # --- Applica la traslazione al Painter per spostare il gradiente ---
                painter.save() # Salva lo stato attuale del painter (posizione, ecc.)
                # Trasla l'origine del painter all'inizio del chunk, più lo spostamento animato
                # Il gradiente definito da (0,0) a (50,0) verrà disegnato a partire da questa nuova origine
                painter.translate(chunkFillRect.left() - offset, chunkFillRect.top())

                # --- Disegna il rettangolo riempito nel sistema di coordinate traslato ---
                # Disegniamo un rettangolo la cui larghezza è la larghezza del chunkFillRect
                # e altezza è l'altezza del chunkFillRect, partendo da (0,0) nel sistema traslato.
                painter.setBrush(QBrush(gradient)) # Usa il gradiente come riempimento
                painter.setPen(Qt.NoPen) # Nessun bordo per il chunk

                # Disegna il rettangolo riempito. Il gradiente si muoverà a causa della traslazione animata.
                # La larghezza qui è la larghezza del chunk effettivo.
                painter.drawRect(QRectF(0, 0, chunkFillRect.width() + offset, chunkFillRect.height())) # Aggiusta larghezza per coprire l'offset iniziale?


                painter.restore() # Ripristina lo stato del painter (rimuove la traslazione)

                # --- Esempio di Disegno di una Linea Mobile Sopra (Opzionale) ---
                # Se vuoi disegnare qualcosa SOPRA il gradiente animato, fallo qui DOPO painter.restore()
                # Ad esempio, la linea bianca semi-trasparente che si muove
                # lineX_absolute = chunkFillRect.left() + (self._animation_phase / 199.0) * chunkFillRect.width()
                # painter.setPen(QPen(QColor(255, 255, 255, 100), 2))
                # painter.drawLine(int(lineX_absolute), int(chunkFillRect.top()), int(lineX_absolute), int(chunkFillRect.bottom()))


# --- Fine Classe CustomProgressBar (Metodo Aggiornato) ---


# --- Classe App (Rimane invariata) ---
class App(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Controllo Stato Richiesta - NSIS")
        self.setGeometry(100, 100, 1200, 800)

        main_layout = QHBoxLayout(self)

        self.view = QWebEngineView()
        self.view.setPage(WebEnginePage(self.view))
        main_layout.addWidget(self.view, stretch=3)

        right_column_layout = QVBoxLayout()
        main_layout.addLayout(right_column_layout, stretch=1)

        ctrl_container = QFrame()
        ctrl_container.setStyleSheet("""
            QFrame {
                background-color: #f8f8f8;
                border: 1px solid #dddddd;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        ctrl = QVBoxLayout(ctrl_container)
        ctrl.setContentsMargins(5, 5, 5, 5)
        ctrl.setSpacing(8)

        right_column_layout.addWidget(ctrl_container, stretch=1)

        professional_button_style = """
            QPushButton {
                background-color: #34495e;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #2c3e50;
            }
            QPushButton:pressed {
                background-color: #1a242f;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """

        self.btn_open = QPushButton("Apri NSIS")
        self.btn_open.clicked.connect(lambda: self.view.load(
            QUrl('https://www.impresa.gov.it/intro/info/news.html')
        ))
        self.btn_open.setStyleSheet(professional_button_style)
        ctrl.addWidget(self.btn_open)

        self.btn_start = QPushButton("Seleziona file e Avvia")
        self.btn_start.clicked.connect(self.start)
        self.btn_start.setStyleSheet(professional_button_style)
        ctrl.addWidget(self.btn_start)

        self.btn_stop = QPushButton("Interrompi")
        self.btn_stop.clicked.connect(self.stop)
        self.btn_stop.setEnabled(False)
        self.btn_stop.setStyleSheet(professional_button_style)
        ctrl.addWidget(self.btn_stop)

        ctrl.addSpacing(10)
        line1 = QFrame()
        line1.setFrameShape(QFrame.HLine)
        line1.setFrameShadow(QFrame.Sunken)
        ctrl.addWidget(line1)
        ctrl.addSpacing(10)

        # --- Progress bar con Label separata ---
        progress_layout = QHBoxLayout()
        progress_layout.setContentsMargins(0,0,0,0)
        progress_layout.setSpacing(5)

        # ** Usa la tua CustomProgressBar qui **
        self.progress = CustomProgressBar() # *** Usiamo la classe personalizzata ***
        self.progress.setRange(0, 100) # Imposta un range di default

        progress_layout.addWidget(self.progress, stretch=1)

        self.progress_label = QLabel("0%") # Label per il percentuale
        self.progress_label.setStyleSheet("""
            QLabel {
                color: #333333;
                font-size: 10px;
                min-width: 30px;
                max-width: 40px;
            }
        """)
        self.progress_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        progress_layout.addWidget(self.progress_label)

        ctrl.addLayout(progress_layout)
        # --- Fine Progress bar con Label separata ---


        self.status = QLabel("", alignment=Qt.AlignCenter)
        ctrl.addWidget(self.status)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setStyleSheet("""
            QTextEdit {
                border: 1px solid #cccccc;
                border-radius: 8px;
                padding: 10px;
                box-shadow: 2px 2px 5px rgba(0, 0, 0, 0.1);
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #ffffff, stop: 1 #f0f0f0);
                font-family: Consolas, Monaco, 'Andale Mono', 'Ubuntu Mono', monospace;
                font-size: 10px;
            }
        """)
        ctrl.addWidget(self.log)

        ctrl.addSpacing(10)
        line2 = QFrame()
        line2.setFrameShape(QFrame.HLine)
        line2.setFrameShadow(QFrame.Sunken)
        ctrl.addWidget(line2)
        ctrl.addSpacing(10)

        self.badge_layout = QVBoxLayout()
        ctrl.addLayout(self.badge_layout)
        self._badge_widgets = set()

        subdued_color = "#bdc3c7"
        vibrant_closed_color = "#27ae60"

        self.card_annullata, self.label_annullata, self.sparkle_annullata = self.create_badge("🟡", "Annullate", subdued_color)
        self.card_aperta, self.label_aperta, self.sparkle_aperta = self.create_badge("🟢", "Aperte", subdued_color)
        self.card_chiusa, self.label_chiusa, self.sparkle_chiusa = self.create_badge("✅", "Chiuse", vibrant_closed_color)
        self.card_lavorazione, self.label_lavorazione, self.sparkle_lavorazione = self.create_badge("🟠", "In lavorazione", subdued_color)
        self.card_inviata, self.label_inviata, self.sparkle_inviata = self.create_badge("📤", "Inviate", subdued_color)
        self.card_eccezioni, self.label_eccezioni, self.sparkle_eccezioni = self.create_badge("❗", "Eccezioni", subdued_color)

        self.card_annullata.setVisible(False)
        self.card_aperta.setVisible(False)
        self.card_chiusa.setVisible(False)
        self.card_lavorazione.setVisible(False)
        self.card_inviata.setVisible(False)
        self.card_eccezioni.setVisible(False)

        right_column_layout.addStretch()

        firma_h_layout = QHBoxLayout()
        firma_h_layout.setContentsMargins(0, 0, 0, 0)
        firma_h_layout.addStretch()
        self.firma_label = QLabel("<i>©2025 ST, version 1.0</i>")
        self.firma_label.setStyleSheet("""
            QLabel {
                color: #777777;
                font-size: 10px;
                margin-right: 15px;
            }
        """)
        firma_h_layout.addWidget(self.firma_label)
        right_column_layout.addLayout(firma_h_layout)

        self.bridge = JSBridge()
        self.channel = QWebChannel()
        self.channel.registerObject('bridge', self.bridge)
        self.view.page().setWebChannel(self.channel)

        self.max_retries = 2
        self._stop = False

    def create_badge(self, icon, label_text, color):
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 6px;
                padding: 2px 4px;
                border: none;
            }}
            QLabel {{
                color: black;
                font-weight: 600;
                font-size: 11px;
            }}
        """)
        frame.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Fixed)

        layout = QVBoxLayout(frame)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(1)
        layout.setAlignment(Qt.AlignCenter)

        label = QLabel(f"{icon} {label_text}: 0")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        sparkle = QLabel("")
        sparkle.setAlignment(Qt.AlignCenter)
        sparkle.setFixedHeight(12)

        layout.addWidget(label)
        layout.addWidget(sparkle)

        return frame, label, sparkle

    def flash_emoji(self, sparkle_label, emoji="✨"):
        sparkle_label.setText(emoji)
        QTimer.singleShot(700, lambda: sparkle_label.setText(""))

    def animate_badge(self, label):
        effect = QGraphicsOpacityEffect()
        label.setGraphicsEffect(effect)

        animation = QPropertyAnimation(effect, b"opacity")
        animation.setStartValue(0.3)
        animation.setEndValue(1.0)
        animation.setDuration(600)
        animation.start()

    def start(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Seleziona file Excel", "", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self._stop = False
        self.status.setText("⏳ Avvio automazione...")
        self.status.setStyleSheet("color: #e67e22;")

        self.log.clear()
        self.log_message("🔄 Avvio elaborazione codici...")

        self.label_annullata.setText("🟡 Annullate: 0")
        self.label_aperta.setText("🟢 Aperte: 0")
        self.label_chiusa.setText("✅ Chiuse: 0")
        self.label_lavorazione.setText("🟠 In lavorazione: 0")
        self.label_inviata.setText("📤 Inviate: 0")
        self.label_eccezioni.setText("❗ Eccezioni: 0")

        try:
            df = pd.read_excel(file_path)
            if "Ricerca" not in df.columns:
                self.status.setText("❌ Errore: la colonna 'Ricerca' non è presente nel file.")
                self.status.setStyleSheet("color: #e74c3c;")
                self.btn_start.setEnabled(True)
                self.btn_stop.setEnabled(False)
                return
            codes = df['Ricerca'].astype(str).tolist()
        except Exception as e:
            self.status.setText(f"❌ Errore durante la lettura del file: {e}")
            self.status.setStyleSheet("color: #e74c3c;")
            self.btn_start.setEnabled(True)
            self.btn_stop.setEnabled(False)
            return

        self.progress.setMaximum(len(codes))
        self.progress.setValue(0)
        self.progress_label.setText("0%")
        self.results = []

        self.run_checks(codes, file_path)

    def stop(self):
        self._stop = True
        self.status.setText("❌ Interrotto dall'utente")
        self.status.setStyleSheet("color: #e74c3c;")

    def log_message(self, message):
        self.log.append(message)
        self.log.verticalScrollBar().setValue(self.log.verticalScrollBar().maximum())

    def fetch_state(self, code):
        for attempt in range(self.max_retries + 1):
            if attempt > 0:
                self.status.setText(f"⏳ Riprovo '{code}' (tentativo {attempt}/{self.max_retries})")
                self.status.setStyleSheet("color: #e67e22;")
                time.sleep(1)

            js_input = f"document.getElementById('codiceRichiesta').value='{code}';"
            js_click = "document.getElementById('cercaRichiestaNullaOstaBtn').click();"

            self.view.page().runJavaScript(js_input)
            time.sleep(0.1)
            self.view.page().runJavaScript(js_click)

            for _ in range(100):
                cells = self.bridge.evaluate(self.view.page(), ALL_CELLS_JS)
                if cells and len(cells) > 8 and str(cells[8]) == str(code).strip():
                    state = cells[2].strip() if len(cells) > 2 else 'Sconosciuto'
                    note = cells[10].strip() if len(cells) > 10 and cells[10] is not None else ''
                    self.last_cells = [c.strip() if c is not None else '' for c in cells]
                    return state
                time.sleep(0.1)

        self.log_message(f"⚠️ Timeout o nessun risultato valido trovato per il codice: {code}")
        self.last_cells = [''] * 11
        self.last_cells[2] = 'Errore o Nessun risultato'
        self.last_cells[8] = str(code).strip()
        return 'Errore o Nessun risultato'

    def run_checks(self, codes, file_path):
        count_annullata = 0
        count_aperta = 0
        count_chiusa = 0
        count_lavorazione = 0
        count_inviata = 0
        count_eccezioni = 0

        for idx, code in enumerate(codes, start=1):
            if self._stop:
                break

            self.status.setText(f"⏳ Processando: {code} ({idx}/{len(codes)})")
            self.status.setStyleSheet("color: #e67e22;")

            cleaned_input_code = str(code).strip()

            state = self.fetch_state(cleaned_input_code)
            self.log_message(f"{idx}/{len(codes)} ➜ Codice: {cleaned_input_code} | Stato: {state}")

            normalized_state = state.strip().upper()

            if normalized_state == "ANNULLATA":
                count_annullata += 1
            elif normalized_state == "APERTA":
                count_aperta += 1
            elif normalized_state == "CHIUSA":
                count_chiusa += 1
            elif normalized_state == "IN LAVORAZIONE":
                count_lavorazione += 1
            elif normalized_state == "INVIATA":
                count_inviata += 1
            else:
                count_eccezioni += 1

            def add_badge_if_needed(card):
                if card not in self._badge_widgets:
                    self.badge_layout.addWidget(card)
                    self._badge_widgets.add(card)
                card.setVisible(True)

            old_text = self.label_annullata.text()
            new_text = f"🟡 Annullate: {count_annullata}"
            if old_text != new_text:
                self.label_annullata.setText(new_text)
                self.flash_emoji(self.sparkle_annullata)
                if count_annullata > 0: add_badge_if_needed(self.card_annullata)

            old_text = self.label_aperta.text()
            new_text = f"🟢 Aperte: {count_aperta}"
            if old_text != new_text:
                self.label_aperta.setText(new_text)
                self.flash_emoji(self.sparkle_aperta)
                if count_aperta > 0: add_badge_if_needed(self.card_aperta)

            old_text = self.label_chiusa.text()
            new_text = f"✅ Chiuse: {count_chiusa}"
            if old_text != new_text:
                self.label_chiusa.setText(new_text)
                self.flash_emoji(self.sparkle_chiusa)
                if count_chiusa > 0: add_badge_if_needed(self.card_chiusa)

            old_text = self.label_lavorazione.text()
            new_text = f"🟠 In lavorazione: {count_lavorazione}"
            if old_text != new_text:
                self.label_lavorazione.setText(new_text)
                self.flash_emoji(self.sparkle_lavorazione)
                if count_lavorazione > 0: add_badge_if_needed(self.card_lavorazione)

            old_text = self.label_inviata.text()
            new_text = f"📤 Inviate: {count_inviata}"
            if old_text != new_text:
                self.label_inviata.setText(new_text)
                self.flash_emoji(self.sparkle_inviata)
                if count_inviata > 0: add_badge_if_needed(self.card_inviata)

            old_text = self.label_eccezioni.text()
            new_text = f"❗ Eccezioni: {count_eccezioni}"
            if old_text != new_text:
                self.label_eccezioni.setText(new_text)
                self.flash_emoji(self.sparkle_eccezioni)
                if count_eccezioni > 0: add_badge_if_needed(self.card_eccezioni)

            cells = self.last_cells
            stato_res = cells[2] if len(cells) > 2 else ''
            protocollo_uscita_res = cells[5] if len(cells) > 5 else ''
            provvedimento_res = cells[6] if len(cells) > 6 else ''
            data_provvedimento_res = cells[7] if len(cells) > 7 else ''
            codice_richiesta_risultato_res = cells[8] if len(cells) > 8 else ''
            note_usmaf_res = cells[10] if len(cells) > 10 else ''

            self.results.append({
                'Input Code': cleaned_input_code,
                'Stato': stato_res,
                'Protocollo uscita': protocollo_uscita_res,
                'Provvedimento': provvedimento_res,
                'Data Provvedimento': data_provvedimento_res,
                'Codice richiesta (risultato)': codice_richiesta_risultato_res,
                'Note Usmaf': note_usmaf_res
            })

            self.progress.setValue(idx)
            if self.progress.maximum() > self.progress.minimum():
                 percentage = int(((self.progress.value() - self.progress.minimum()) / (self.progress.maximum() - self.progress.minimum())) * 100)
            else:
                 percentage = 0
            self.progress_label.setText(f"{percentage}%")


        if self._stop:
            self.status.setText("❌ Elaborazione interrotta.")
            self.status.setStyleSheet("color: #e74c3c;")
            if self.progress.maximum() > self.progress.minimum():
                 percentage = int(((self.progress.value() - self.progress.minimum()) / (self.progress.maximum() - self.progress.minimum())) * 100)
                 self.progress_label.setText(f"{percentage}%")
            else:
                 self.progress_label.setText("0%")

        else:
            self.status.setText("⏳ Salvataggio su file Excel...")
            self.status.setStyleSheet("color: #e67e22;")
            try:
                wb = load_workbook(file_path)
                ws = wb.active

                for res in self.results:
                    original_row_index = None
                    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=2, max_col=2), start=2):
                         cell_value = str(row[0].value).strip() if row[0].value is not None else ""
                         if cell_value == res['Input Code']:
                            original_row_index = r_idx
                            break

                    if original_row_index is not None:
                        ws.cell(row=original_row_index, column=3, value=res['Stato'])
                        ws.cell(row=original_row_index, column=4, value=res['Protocollo uscita'])
                        ws.cell(row=original_row_index, column=5, value=res['Provvedimento'])
                        ws.cell(row=original_row_index, column=6, value=res['Data Provvedimento'])
                        ws.cell(row=original_row_index, column=7, value=res['Codice richiesta (risultato)'])
                        ws.cell(row=original_row_index, column=8, value=res['Note Usmaf'])
                    else:
                         self.log_message(f"⚠️ Attenzione: impossibile trovare riga originale nel file Excel (Colonna B) per codice input: {res['Input Code']}. Dati non salvati per questo codice.")

                for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                     cell = row[0]
                     cell.number_format = numbers.FORMAT_TEXT

                wb.save(file_path)
                self.status.setText('✅ Completato!')
                self.status.setStyleSheet("color: #2ecc71;")
                self.log_message("✅ Elaborazione completata e file salvato.")
                self.progress_label.setText("100%")

            except Exception as e:
                self.status.setText(f"❌ Errore durante il salvataggio del file: {e}")
                self.status.setStyleSheet("color: #e74c3c;")
                self.log_message(f"❌ Errore durante il salvataggio: {e}")

        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._stop = False


if __name__ == '__main__':
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    window = App()
    window.show()
    sys.exit(app.exec_())