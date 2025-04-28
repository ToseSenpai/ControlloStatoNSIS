# -*- coding: utf-8 -*-
# main_window.py
# Contains the main application window class and the Worker class
# Versione aggiornata con stile Luma (Fase 1) e correzioni layout/stile

import os
import sys
import time
import datetime
import logging
import logging.handlers
from PyQt6 import QtCore, QtGui, QtWidgets, QtWebEngineWidgets, QtWebChannel, QtWebEngineCore

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, Color, PatternFill, numbers, Alignment
from openpyxl.utils import get_column_letter
import traceback

# Importa qtawesome e imposta flag di disponibilità
try:
    import qtawesome as qta
    QTAWESOME_AVAILABLE = True
    print("INFO: Libreria 'qtawesome' caricata correttamente.")
except ImportError:
    QTAWESOME_AVAILABLE = False
    print("WARNING: Libreria 'qtawesome' non trovata. Icone standard verranno usate.")

# Import from our modules
from config import (
    MAX_RETRIES, FETCH_TIMEOUT_MS, FETCH_CHECK_INTERVAL_MS, STATO_SELECTOR,
    ALL_CELLS_JS, DELAY_AFTER_INPUT_JS, DELAY_AFTER_CLICK_JS, DELAY_BETWEEN_RETRIES,
    MAX_NULL_CHECKS, URL_NSIS,
    COL_RICERCA, COL_STATO, COL_PROTOCOLLO, COL_PROVVEDIMENTO,
    COL_DATA_PROVV, COL_CODICE_RIS, COL_NOTE,
    # Import colori e stili Luma
    COLOR_LUMA_WHITE, COLOR_LUMA_GRAY_10, COLOR_LUMA_GRAY_30, COLOR_LUMA_GRAY_50,
    COLOR_LUMA_GRAY_70, COLOR_LUMA_GRAY_90, COLOR_LUMA_PURPLE_400, COLOR_LUMA_PURPLE_500,
    LUMA_BORDER_RADIUS_BUTTON, LUMA_BORDER_RADIUS_CONTAINER, LUMA_BORDER_COLOR_INPUT,
    # Import colori badge ridefiniti
    COLOR_ANNULATA, COLOR_APERTA, COLOR_CHIUSA, COLOR_LAVORAZIONE, COLOR_INVIATA, COLOR_ECCEZIONI
)

# Assicurati che questi import esistano o usa i fallbacks
try:
    from ui_components import CustomProgressBar, SpinnerWidget
    from web_engine import WebEnginePage, JSBridge
except ImportError as e:
    logging.basicConfig(level=logging.ERROR) # Setup logger base se fallisce qui
    logging.error(f"Impossibile importare ui_components o web_engine: {e}", exc_info=True)
    # Basic Fallbacks
    class CustomProgressBar(QtWidgets.QProgressBar): pass
    class SpinnerWidget(QtWidgets.QWidget):
        def __init__(self, parent=None): super().__init__(parent); self.hide()
        def startAnimation(self): self.show()
        def stopAnimation(self): self.hide()
        def setColor(self, color): pass # No-op
    class WebEnginePage(QtWebEngineWidgets.QWebEnginePage): pass
    class JSBridge(QtCore.QObject): pass

# ------------------- Worker Class (per QThread) -------------------
class Worker(QtCore.QObject):
    progress = QtCore.pyqtSignal(int, int)
    statusUpdate = QtCore.pyqtSignal(str)
    logUpdate = QtCore.pyqtSignal(str)
    badgeUpdate = QtCore.pyqtSignal(str, int)
    finished = QtCore.pyqtSignal()
    resultsReady = QtCore.pyqtSignal(list)
    requestFetch = QtCore.pyqtSignal(str)

    def __init__(self, codes_to_process):
        super().__init__()
        self.codes = codes_to_process
        self._stop_requested = False
        self._results = []
        self._current_code_index = 0
        self._total_codes = len(codes_to_process)
        self._counts = {"annullata": 0, "aperta": 0, "chiusa": 0, "lavorazione": 0, "inviata": 0, "eccezioni": 0}
        self._badge_map = {
            "ANNULLATA": "🟡 Annullate", "APERTA": "🟢 Aperte", "CHIUSA": "✅ Chiuse",
            "IN LAVORAZIONE": "🟠 In lavorazione", "INVIATA": "📤 Inviate",
            "ECCEZIONE": "❗ Eccezioni"
        }
        self._count_keys = {
            "ANNULLATA": "annullata", "APERTA": "aperta", "CHIUSA": "chiusa",
            "IN LAVORAZIONE": "lavorazione", "INVIATA": "inviata"
        }
        self._exception_states = {
            "NON TROVATO", "ERRORE TIMEOUT", "ERRORE PAGINA", "ERRORE INTERNO FETCH",
            "SCONOSCIUTO", "INTERROTTO",
            "ERRORE CARICAMENTO PAGINA", "ELEMENTO NON TROVATO (JS)", "ERRORE RISULTATO JS"
        }
        self._finished_emitted_after_stop = False

    def run(self):
        """Main execution logic for the worker thread."""
        print("--- Worker Thread Started ---")
        self._results = []; self._counts = {k: 0 for k in self._counts}; self._stop_requested = False
        self._finished_emitted_after_stop = False
        if not self.codes:
            self.logUpdate.emit("Nessun codice da processare nel worker.")
            self.finished.emit(); return

        self.progress.emit(0, self._total_codes)
        self._current_code_index = 0

        if not self._stop_requested:
            if self._current_code_index < len(self.codes):
                code = self.codes[self._current_code_index]
                self.statusUpdate.emit(f"Inizio elaborazione per: {code} (1/{self._total_codes})")
                self.requestFetch.emit(code)
            else:
                self.logUpdate.emit("Lista codici vuota."); self.finished.emit()
        else:
            self.logUpdate.emit("❌ Interrotto prima dell'inizio."); self.finished.emit()

    @QtCore.pyqtSlot(str, str, list)
    def processFetchedResult(self, code, state, last_cells):
        """Processes the result received from the main thread for a code."""
        if self._stop_requested:
            if not self._finished_emitted_after_stop:
                thread = self.thread()
                if thread and thread.isRunning():
                    self.logUpdate.emit(f"❌ Elaborazione interrotta nel worker (risultato per {code} ricevuto: {state}). Fine.")
                    self.resultsReady.emit(self._results)
                    self.finished.emit()
                    self._finished_emitted_after_stop = True
            return

        current_idx = self._current_code_index
        self.logUpdate.emit(f"{current_idx + 1}/{self._total_codes} ➜ Codice: {code} | Stato Web: {state}")

        normalized_state_upper = state.strip().upper()
        if normalized_state_upper in self._exception_states:
            status_key_for_count = "eccezioni"
            badge_prefix = self._badge_map["ECCEZIONE"]
        else:
            status_key_for_count = self._count_keys.get(normalized_state_upper, "eccezioni")
            badge_prefix = self._badge_map.get(normalized_state_upper, self._badge_map["ECCEZIONE"])
            if status_key_for_count == "eccezioni" and normalized_state_upper not in self._exception_states:
                 self.logUpdate.emit(f"⚠️ Stato non mappato: '{state}'. Conteggiato come Eccezione.")

        if status_key_for_count not in self._counts:
             self.logUpdate.emit(f"⚠️ Chiave conteggio interna non prevista: {status_key_for_count} per stato {state}")
             status_key_for_count = "eccezioni"

        self._counts[status_key_for_count] += 1
        count_to_display = self._counts[status_key_for_count]
        self.badgeUpdate.emit(badge_prefix, count_to_display)

        stato_res = state
        protocollo_uscita_res = last_cells[5] if len(last_cells) > 5 else ''
        provvedimento_res = last_cells[6] if len(last_cells) > 6 else ''
        data_provvedimento_res = last_cells[7] if len(last_cells) > 7 else ''
        codice_richiesta_risultato_res = last_cells[8] if len(last_cells) > 8 and last_cells[8] else code
        note_usmaf_res = last_cells[10] if len(last_cells) > 10 else ''
        if normalized_state_upper in self._exception_states:
            note_usmaf_res = f"Stato recuperato: {state}. {note_usmaf_res}".strip()

        self._results.append({
            'Input Code': code, 'Stato': stato_res, 'Protocollo uscita': protocollo_uscita_res,
            'Provvedimento': provvedimento_res, 'Data Provvedimento': data_provvedimento_res,
            'Codice richiesta (risultato)': codice_richiesta_risultato_res, 'Note Usmaf': note_usmaf_res
        })

        self.progress.emit(current_idx + 1, self._total_codes)
        self._current_code_index += 1

        if self._current_code_index < self._total_codes:
            next_code = self.codes[self._current_code_index]
            self.statusUpdate.emit(f"Richiesta fetch per: {next_code} ({self._current_code_index + 1}/{self._total_codes})")
            self.requestFetch.emit(next_code)
        else:
            self.logUpdate.emit("✅ Elaborazione codici completata.")
            self.statusUpdate.emit("✅ Elaborazione completata.")
            self.resultsReady.emit(self._results)
            self.finished.emit()

    def request_stop(self):
        """Sets the stop flag for the worker."""
        print("--- Worker Stop Requested ---")
        self._stop_requested = True
# ------------------- Fine Worker Class -------------------


# ------------------- Classe App Principale -------------------
logger = logging.getLogger(__name__)

class App(QtWidgets.QWidget):
    """Classe principale dell'applicazione GUI."""
    fetchCompleted = QtCore.pyqtSignal(str, str, list)

    def __init__(self, ui_font_family='Arial'): # Accetta il font
        super().__init__()
        self.setWindowTitle("Controllo Stato Richiesta - NSIS")
        self.setGeometry(QtCore.QRect(100, 100, 1200, 800))
        self.ui_font_family = ui_font_family # Memorizza font

        # Carica icona app
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            icon_path = os.path.join(script_dir, "icon.ico")
            if os.path.exists(icon_path): self.setWindowIcon(QtGui.QIcon(icon_path)); logger.info("Icona applicazione caricata.")
            else: logger.warning(f"File icona non trovato: {icon_path}")
        except Exception as e: logger.error(f"Errore caricamento icona: {e}", exc_info=True)

        # Inizializza variabili membro
        self._badge_widgets = set()
        self._badge_widgets_map = {}
        self.max_retries = MAX_RETRIES
        self.last_cells = []
        self.current_file_path = None
        self.thread = None
        self.worker = None
        self._current_fetch_context = None
        self._active_badge_animations = {}

        # Colori icone (Usa colori Luma definiti in config)
        self.icon_color_nav = COLOR_LUMA_GRAY_50
        self.icon_color_open = COLOR_LUMA_GRAY_90
        self.icon_color_start = COLOR_LUMA_GRAY_90
        self.icon_color_stop = COLOR_ECCEZIONI # Usa rosso eccezioni per icona stop

        # Setup iniziale
        self._setup_logging()
        self._setup_ui() # Chiamato DOPO aver definito self.ui_font_family
        self._setup_webengine()

        # Stylesheet Globale Base (Applica font, colore testo base)
        self.setStyleSheet(f"""
            QWidget {{
                font-family: "{self.ui_font_family}", Arial, sans-serif;
                font-size: 12px;
                background-color: {COLOR_LUMA_WHITE}; /* Sfondo base */
                color: {COLOR_LUMA_GRAY_90}; /* Testo base */
            }}
            /* Stili più specifici verranno applicati nei rispettivi widget/container */
        """)

        logger.info(f"Applicazione inizializzata. Font UI: '{self.ui_font_family}'")

    @QtCore.pyqtSlot()
    def _clear_log(self):
        """Slot per pulire il contenuto del widget di log."""
        user_confirmation = QtWidgets.QMessageBox.question(self, "Conferma Pulizia Log",
                                                           "Sei sicuro di voler cancellare tutto il contenuto del log?",
                                                           QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                                                           QtWidgets.QMessageBox.StandardButton.No)
        if user_confirmation == QtWidgets.QMessageBox.StandardButton.Yes:
            if hasattr(self, 'log') and self.log:
                self.log.clear()
                logger.info("Log pulito dall'utente.")
                self.log.append("<i>Log pulito.</i>")

    def _setup_logging(self):
        """Configura il modulo logging per Console e File."""
        log_level = logging.DEBUG
        log_format_standard = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(name)s - %(message)s')
        log_file = "app_log.log"

        root_logger = logging.getLogger()
        if not root_logger.hasHandlers():
             root_logger.setLevel(log_level)
             for handler in root_logger.handlers[:]:
                 root_logger.removeHandler(handler)
                 handler.close()

             console_handler = logging.StreamHandler(sys.stdout)
             console_handler.setFormatter(log_format_standard)
             console_handler.setLevel(logging.INFO)
             root_logger.addHandler(console_handler)

             try:
                 file_handler = logging.handlers.RotatingFileHandler(
                     log_file, maxBytes=1 * 1024 * 1024, backupCount=3, encoding='utf-8'
                 )
                 file_handler.setFormatter(log_format_standard)
                 file_handler.setLevel(logging.DEBUG)
                 root_logger.addHandler(file_handler)
             except Exception as e:
                  print(f"ERRORE CRITICO: Impossibile configurare FileHandler per logging: {e}")

             logging.getLogger("openpyxl").setLevel(logging.WARNING)
             logging.getLogger("urllib3").setLevel(logging.WARNING)
             logger.info("Sistema di logging configurato (Console=INFO, File=DEBUG).")
        else:
             logger.info("Sistema di logging già configurato.")

    # Inserisci questo metodo aggiornato nella classe App
    # sostituendo il vecchio _setup_ui

    def _setup_ui(self):
        """Crea e organizza i widget dell'interfaccia utente con stile Luma/DHL."""
        main_layout = QtWidgets.QHBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)  # Margine finestra
        main_layout.setSpacing(10)  # Spazio tra WebView e Pannello Controlli

        # --- Web View Frame (Aggiornato stile) ---
        web_view_frame = QtWidgets.QFrame()
        web_view_frame.setObjectName("WebViewContainer")
        web_view_frame.setStyleSheet(f"""
            QFrame#WebViewContainer {{
                border: 1px solid {COLOR_LUMA_GRAY_30}; /* Bordo Luma */
                border-radius: {LUMA_BORDER_RADIUS_CONTAINER}; /* Arrotondamento Luma */
                background-color: {COLOR_LUMA_WHITE};
            }}
        """)
        web_view_frame.setContentsMargins(1, 1, 1, 1)  # Padding minimo interno al bordo

        web_view_layout = QtWidgets.QVBoxLayout(web_view_frame)
        web_view_layout.setContentsMargins(8, 8, 8, 8)  # Padding interno frame
        web_view_layout.setSpacing(6)

        # --- CREA LA WEB VIEW PRIMA DELLA TOOLBAR NAV ---
        self.view = QtWebEngineWidgets.QWebEngineView()
        # Puoi già impostare proprietà della vista qui se necessario
        # Esempio: self.view.settings().setAttribute(...)

        # --- Barra Navigazione Web (Aggiornato stile bottoni) ---
        nav_toolbar_layout = QtWidgets.QHBoxLayout()
        nav_toolbar_layout.setContentsMargins(0, 0, 0, 0)
        nav_toolbar_layout.setSpacing(4)
        nav_button_style = f""" QPushButton {{ background-color: {COLOR_LUMA_WHITE}; border: 1px solid {COLOR_LUMA_GRAY_30}; border-radius: {LUMA_BORDER_RADIUS_BUTTON}; padding: 4px; min-width: 26px; max-width: 26px; min-height: 26px; max-height: 26px; outline: none; }} QPushButton:hover {{ background-color: {COLOR_LUMA_GRAY_10}; }} QPushButton:pressed {{ background-color: {COLOR_LUMA_GRAY_30}; }} QPushButton:disabled {{ background-color: {COLOR_LUMA_GRAY_10}; border-color: {COLOR_LUMA_GRAY_30}; }} """

        self.btn_back = QtWidgets.QPushButton()
        if QTAWESOME_AVAILABLE:
            icon_back = qta.icon('fa5s.arrow-left', color=self.icon_color_nav, color_disabled=COLOR_LUMA_GRAY_50)
        else:
            icon_back = self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_ArrowBack)
        self.btn_back.setIcon(icon_back);
        self.btn_back.setToolTip("Indietro")
        self.btn_back.setStyleSheet(nav_button_style);
        self.btn_back.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.btn_back.setEnabled(False);
        self.btn_back.clicked.connect(self.view.back)  # Connesso a self.view
        nav_toolbar_layout.addWidget(self.btn_back)

        self.btn_forward = QtWidgets.QPushButton()
        if QTAWESOME_AVAILABLE:
            icon_forward = qta.icon('fa5s.arrow-right', color=self.icon_color_nav, color_disabled=COLOR_LUMA_GRAY_50)
        else:
            icon_forward = self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_ArrowForward)
        self.btn_forward.setIcon(icon_forward);
        self.btn_forward.setToolTip("Avanti")
        self.btn_forward.setStyleSheet(nav_button_style);
        self.btn_forward.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.btn_forward.setEnabled(False);
        self.btn_forward.clicked.connect(self.view.forward)  # Connesso a self.view
        nav_toolbar_layout.addWidget(self.btn_forward)

        self.btn_reload = QtWidgets.QPushButton()
        if QTAWESOME_AVAILABLE:
            icon_reload = qta.icon('fa5s.sync-alt', color=self.icon_color_nav, color_disabled=COLOR_LUMA_GRAY_50)
        else:
            icon_reload = self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_BrowserReload)
        self.btn_reload.setIcon(icon_reload);
        self.btn_reload.setToolTip("Ricarica")
        self.btn_reload.setStyleSheet(nav_button_style);
        self.btn_reload.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.btn_reload.setEnabled(False);
        self.btn_reload.clicked.connect(self.view.reload)  # Connesso a self.view
        nav_toolbar_layout.addWidget(self.btn_reload)

        nav_toolbar_layout.addStretch()
        web_view_layout.addLayout(nav_toolbar_layout)

        # Aggiungi la web view (già creata) al layout SOTTO la toolbar
        web_view_layout.addWidget(self.view, stretch=1)

        main_layout.addWidget(web_view_frame, stretch=3)

        # --- Pannello Controlli Destro ---
        right_column_layout = QtWidgets.QVBoxLayout()
        main_layout.addLayout(right_column_layout, stretch=1)

        ctrl_container = QtWidgets.QFrame()
        ctrl_container.setObjectName("ControlContainer")
        # Stylesheet per il pannello e i suoi widget figli
        ctrl_container.setStyleSheet(f"""
            QFrame#ControlContainer {{
                background-color: {COLOR_LUMA_GRAY_10};
                border: 1px solid {COLOR_LUMA_GRAY_30};
                border-radius: {LUMA_BORDER_RADIUS_CONTAINER};
                padding: 12px;
            }}
            QGroupBox {{
                font-weight: 600;
                border: none;
                padding-top: 18px; /* Spazio sopra per titolo */
                margin-bottom: 15px; /* Spazio sotto gruppo */
                background-color: transparent;
            }}
            QGroupBox::title {{
                subcontrol-origin: padding; /* Posizione titolo legata a padding */
                subcontrol-position: top left;
                left: 10px; padding: 2px 6px;
                margin-bottom: 5px; /* Spazio sotto il titolo */
                color: {COLOR_LUMA_GRAY_90};
                font-size: 12px; font-weight: 600;
                background-color: {COLOR_LUMA_GRAY_30}; /* Sfondo grigio titolo */
                border: none;
                border-radius: 6px;
                min-height: 20px;
            }}
            QLabel {{ /* Default label style nel pannello */
                 background-color: transparent;
                 padding: 1px;
                 color: {COLOR_LUMA_GRAY_70}; /* Colore testo secondario */
            }}
            CustomProgressBar {{
                 border: 1px solid {COLOR_LUMA_GRAY_30};
                 border-radius: {LUMA_BORDER_RADIUS_BUTTON};
                 background-color: {COLOR_LUMA_GRAY_30}; /* Sfondo barra 'vuota' */
                 height: 8px; /* Altezza barra */
                 /* Colore chunk gestito da paintEvent */
            }}
             /* Stile specifico per Log Widget via ObjectName */
             QTextEdit#LogWidget {{
                 background-color: {COLOR_LUMA_WHITE};
                 border: 1px solid {LUMA_BORDER_COLOR_INPUT};
                 border-radius: {LUMA_BORDER_RADIUS_BUTTON};
                 padding: 8px;
                 color: {COLOR_LUMA_GRAY_90};
                 font-size: 10px;
             }}
             /* Stile Scrollbar per Log Widget */
             QTextEdit#LogWidget QScrollBar:vertical {{
                 border: none; background: {COLOR_LUMA_GRAY_10}; width: 10px; margin: 0px;
             }}
             QTextEdit#LogWidget QScrollBar::handle:vertical {{
                 background: {COLOR_LUMA_GRAY_50}; min-height: 20px; border-radius: 5px;
             }}
              QTextEdit#LogWidget QScrollBar::handle:vertical:hover {{
                 background: {COLOR_LUMA_GRAY_70};
             }}
             QTextEdit#LogWidget QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                 height: 0px; background: none;
             }}
             QTextEdit#LogWidget QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                 background: none;
             }}
        """)
        ctrl_layout = QtWidgets.QVBoxLayout(ctrl_container)
        ctrl_layout.setContentsMargins(10, 10, 10, 10)  # Padding interno pannello
        ctrl_layout.setSpacing(15)  # Spazio tra elementi principali verticali
        right_column_layout.addWidget(ctrl_container)

        # --- Barra Titolo "Azioni" con Logo ---
        title_logo_layout = QtWidgets.QHBoxLayout()
        title_logo_layout.setContentsMargins(0, 0, 5, 5);
        title_logo_layout.setSpacing(6)
        azioni_label = QtWidgets.QLabel("Azioni")
        # Stile etichetta "Azioni" che simula un titolo groupbox
        azioni_label.setStyleSheet(
            f""" QLabel {{ font-weight: 600; color: {COLOR_LUMA_GRAY_90}; font-size: 12px; background-color: {COLOR_LUMA_GRAY_30}; border: none; border-radius: 6px; margin-left: 5px; padding: 2px 6px; min-height: 20px; qproperty-alignment: 'AlignVCenter | AlignLeft'; }} """)
        title_logo_layout.addWidget(azioni_label)
        title_logo_layout.addStretch(1)  # Spinge logo a destra
        try:  # Aggiunta logo
            script_dir = os.path.dirname(os.path.abspath(__file__))
            logo_path = os.path.join(script_dir, "ilovecustoms.png")
            if os.path.exists(logo_path):
                logo_pixmap = QtGui.QPixmap(logo_path)
                if not logo_pixmap.isNull():
                    scaled_logo = logo_pixmap.scaledToHeight(20, QtCore.Qt.TransformationMode.SmoothTransformation)
                    logo_widget = QtWidgets.QLabel();
                    logo_widget.setPixmap(scaled_logo)
                    logo_widget.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignVCenter)
                    logo_widget.setToolTip("Logo I Love Customs")
                    title_logo_layout.addWidget(logo_widget)
                else:
                    logger.warning(f"Impossibile caricare QPixmap logo: '{logo_path}'.")
            else:
                logger.warning(f"File logo non trovato: '{logo_path}'.")
        except Exception as e_logo:
            logger.error(f"Errore aggiunta logo: {e_logo}", exc_info=True)
        ctrl_layout.addLayout(title_logo_layout)  # Aggiunge la barra titolo
        # --- Fine Barra Titolo ---

        # --- Definizioni Stili Pulsanti Luma/DHL ---
        professional_button_style = f"""QPushButton {{ background-color: {COLOR_LUMA_WHITE}; color: {COLOR_LUMA_GRAY_90}; border: 1px solid {COLOR_LUMA_GRAY_30}; padding: 8px 15px; border-radius: {LUMA_BORDER_RADIUS_BUTTON}; font-weight: 600; outline: none; text-align: center; }} QPushButton:hover {{ background-color: {COLOR_LUMA_GRAY_10}; }} QPushButton:pressed {{ background-color: {COLOR_LUMA_GRAY_30}; }} QPushButton:disabled {{ background-color: {COLOR_LUMA_GRAY_10}; color: {COLOR_LUMA_GRAY_50}; border-color: {COLOR_LUMA_GRAY_30}; }} """

        # Stile Pulsante Primario con colori DHL
        dhl_yellow = "#FFCC00";
        dhl_red = "#D40511"
        dhl_yellow_hover = "#FFD633";
        dhl_red_hover = "#E0202D"
        dhl_yellow_pressed = "#F0C200";
        dhl_red_pressed = "#C0000A"
        text_color_on_dhl = COLOR_LUMA_GRAY_90  # Testo scuro su giallo/rosso

        primary_button_style = f""" QPushButton {{ background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 {dhl_yellow}, stop:1 {dhl_red}); color: {text_color_on_dhl}; border: none; padding: 9px 15px; border-radius: {LUMA_BORDER_RADIUS_BUTTON}; font-weight: 600; outline: none; text-align: center; }} QPushButton:hover {{ background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 {dhl_yellow_hover}, stop:1 {dhl_red_hover}); }} QPushButton:pressed {{ background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 {dhl_yellow_pressed}, stop:1 {dhl_red_pressed}); }} QPushButton:disabled {{ background: {COLOR_LUMA_GRAY_30}; color: {COLOR_LUMA_GRAY_50}; border: none; }} """

        stop_button_style = f""" QPushButton {{ background-color: {COLOR_LUMA_GRAY_10}; color: {COLOR_ECCEZIONI}; border: 1px solid #FECACA; padding: 8px 15px; border-radius: {LUMA_BORDER_RADIUS_BUTTON}; font-weight: 600; outline: none; text-align: center; }} QPushButton:hover {{ background-color: #FEF2F2; border-color: #FCA5A5; }} QPushButton:pressed {{ background-color: #FEE2E2; }} QPushButton:disabled {{ background-color: {COLOR_LUMA_GRAY_10}; color: {COLOR_LUMA_GRAY_50}; border-color: {COLOR_LUMA_GRAY_30}; }} """

        # --- Widget Azioni ---
        action_group_widget = QtWidgets.QWidget()
        action_group_widget.setStyleSheet("background-color: transparent;")  # Sfondo trasparente
        action_layout = QtWidgets.QVBoxLayout(action_group_widget)
        action_layout.setContentsMargins(0, 8, 0, 8)  # Margini sopra/sotto gruppo
        action_layout.setSpacing(8)  # Spazio tra elementi azioni

        # Pulsante Apri NSIS
        self.btn_open = QtWidgets.QPushButton(" Apri NSIS")
        if QTAWESOME_AVAILABLE: self.btn_open.setIcon(qta.icon('fa5s.external-link-alt', color=self.icon_color_open))
        self.btn_open.setStyleSheet(professional_button_style)
        self.btn_open.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btn_open.setToolTip("Apre il sito NSIS nel browser interno")
        self.btn_open.clicked.connect(self.open_nsis_url_test)
        action_layout.addWidget(self.btn_open)

        # Pulsante Avvia (Usa stile DHL aggiornato)
        self.btn_start = QtWidgets.QPushButton(" Seleziona e Avvia")
        if QTAWESOME_AVAILABLE: self.btn_start.setIcon(
            qta.icon('fa5s.play', color=self.icon_color_start))  # Usa icona scura
        self.btn_start.setStyleSheet(primary_button_style)  # Applica stile DHL
        self.btn_start.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btn_start.setToolTip("Seleziona un file Excel e avvia l'elaborazione")
        self.btn_start.clicked.connect(self.start_processing)
        action_layout.addWidget(self.btn_start)

        # Pulsante Interrompi
        self.btn_stop = QtWidgets.QPushButton(" Interrompi")
        if QTAWESOME_AVAILABLE: self.btn_stop.setIcon(qta.icon('fa5s.stop', color=self.icon_color_stop))
        self.btn_stop.setStyleSheet(stop_button_style)
        self.btn_stop.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btn_stop.setToolTip("Interrompe l'elaborazione corrente")
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_processing)
        action_layout.addWidget(self.btn_stop)

        # Etichetta File
        self.file_label = QtWidgets.QLabel("Nessun file selezionato")
        self.file_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.file_label.setStyleSheet(
            f""" QLabel {{ color: {COLOR_LUMA_GRAY_70}; font-size: 10px; padding: 6px; border: 1px solid {LUMA_BORDER_COLOR_INPUT}; border-radius: {LUMA_BORDER_RADIUS_BUTTON}; background-color: {COLOR_LUMA_WHITE}; min-height: 20px; }} """)
        self.file_label.setToolTip("Mostra il file Excel attualmente caricato")
        action_layout.addWidget(self.file_label)

        ctrl_layout.addWidget(action_group_widget)  # Aggiungi contenitore azioni

        # --- Gruppo Progresso ---
        progress_group = QtWidgets.QGroupBox("Progresso Elaborazione")
        progress_layout_main = QtWidgets.QVBoxLayout(progress_group)
        progress_layout_main.setContentsMargins(10, 5, 10, 8);
        progress_layout_main.setSpacing(5)
        progress_bar_layout = QtWidgets.QHBoxLayout();
        progress_bar_layout.setSpacing(8)

        # Colori DHL definiti per chiarezza
        dhl_yellow_color = QtGui.QColor("#FFCC00")
        dhl_red_color = QtGui.QColor("#D40511")

        # Spinner (Impostato a ROSSO DHL)
        self.spinner = SpinnerWidget(self);
        self.spinner.setFixedSize(16, 16);
        self.spinner.setColor(dhl_red_color)  # <--- MODIFICATO colore spinner
        progress_bar_layout.addWidget(self.spinner)

        # Progress Bar (Impostata a GIALLO DHL)
        self.progress = CustomProgressBar()
        self.progress.setChunkColor(dhl_yellow_color)  # <--- MODIFICATO colore barra
        progress_bar_layout.addWidget(self.progress, stretch=1)

        self.progress_label = QtWidgets.QLabel("0%");
        self.progress_label.setMinimumWidth(35);
        self.progress_label.setStyleSheet(f"color: {COLOR_LUMA_GRAY_90}; font-weight: 600;");
        self.progress_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignVCenter)
        progress_bar_layout.addWidget(self.progress_label)
        progress_layout_main.addLayout(progress_bar_layout)
        self.status = QtWidgets.QLabel("Pronto.");
        self.status.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter);
        self.status.setStyleSheet(f"color: {COLOR_LUMA_GRAY_70}; padding-top: 5px; font-size: 10px;")
        progress_layout_main.addWidget(self.status)
        ctrl_layout.addWidget(progress_group)

        # --- Gruppo Log ---
        log_group = QtWidgets.QGroupBox("Log Operazioni")
        log_layout_main = QtWidgets.QVBoxLayout(log_group)
        log_layout_main.setContentsMargins(0, 5, 0, 0);
        log_layout_main.setSpacing(4)
        self.log = QtWidgets.QTextEdit();
        self.log.setObjectName("LogWidget");
        self.log.setReadOnly(True)  # objectName per stile scrollbar
        log_layout_main.addWidget(self.log, stretch=1)
        log_toolbar_layout = QtWidgets.QHBoxLayout();
        log_toolbar_layout.setContentsMargins(5, 4, 5, 5);
        log_toolbar_layout.addStretch()
        self.btn_clear_log = QtWidgets.QPushButton();
        self.btn_clear_log.setToolTip("Pulisci Log")
        btn_clear_style = f""" QPushButton {{ background-color: transparent; border: 1px solid {COLOR_LUMA_GRAY_30}; border-radius: {LUMA_BORDER_RADIUS_BUTTON}; padding: 3px; min-width: 24px; max-width: 24px; min-height: 24px; max-height: 24px; outline: none; }} QPushButton:hover {{ background-color: {COLOR_LUMA_GRAY_10}; border-color: {COLOR_LUMA_GRAY_50}; }} QPushButton:pressed {{ background-color: {COLOR_LUMA_GRAY_30}; }} """
        self.btn_clear_log.setStyleSheet(btn_clear_style)
        if QTAWESOME_AVAILABLE:
            try:
                icon_clear = qta.icon('fa5s.trash-alt', color=COLOR_LUMA_GRAY_50, color_disabled='#AAAAAA')
            except Exception as e:
                logger.warning(f"Errore icona clear: {e}"); icon_clear = QtGui.QIcon()
            self.btn_clear_log.setIcon(icon_clear)
        self.btn_clear_log.setCursor(QtCore.Qt.CursorShape.PointingHandCursor);
        self.btn_clear_log.clicked.connect(self._clear_log)
        log_toolbar_layout.addWidget(self.btn_clear_log)
        log_layout_main.addLayout(log_toolbar_layout)
        ctrl_layout.addWidget(log_group, stretch=1)

        # --- Gruppo Riepilogo Stati ---
        badge_group = QtWidgets.QGroupBox("Riepilogo Stati")
        self.badge_layout = QtWidgets.QVBoxLayout(badge_group)
        self.badge_layout.setContentsMargins(8, 5, 8, 8);
        self.badge_layout.setSpacing(6)
        badge_data = [("🟡", "Annullate", COLOR_ANNULATA), ("🟢", "Aperte", COLOR_APERTA), ("✅", "Chiuse", COLOR_CHIUSA),
                      ("🟠", "In lavorazione", COLOR_LAVORAZIONE), ("📤", "Inviate", COLOR_INVIATA),
                      ("❗", "Eccezioni", COLOR_ECCEZIONI)]
        self._badge_info_list = [];
        self._badge_widgets_map = {}
        for icon, text, color in badge_data:
            badge_tuple = self.create_badge(icon, text, color)  # Usa funzione aggiornata
            self._badge_info_list.append(badge_tuple)
            prefix = badge_tuple[0].property("badgePrefix")
            if prefix: self._badge_widgets_map[prefix] = badge_tuple
            badge_tuple[0].setVisible(False)  # Nascosto all'inizio
        ctrl_layout.addWidget(badge_group)

        # Firma in fondo
        right_column_layout.addStretch()  # Spinge firma in fondo
        firma_h_layout = QtWidgets.QHBoxLayout();
        firma_h_layout.setContentsMargins(0, 5, 5, 0)
        firma_h_layout.addStretch()
        self.firma_label = QtWidgets.QLabel("<i>©2025 ST, version 1.5-stable</i>")  # Aggiorna versione
        self.firma_label.setStyleSheet(
            f"QLabel {{ color: {COLOR_LUMA_GRAY_50}; font-size: 9px; background-color: transparent; }}")
        self.firma_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignVCenter)
        firma_h_layout.addWidget(self.firma_label)
        right_column_layout.addLayout(firma_h_layout)

        # Connessioni Navigazione Web
        self.view.urlChanged.connect(self._update_navigation_buttons_state)
        self.view.loadFinished.connect(self._update_navigation_buttons_state)
        self._update_navigation_buttons_state()  # Imposta stato iniziale bottoni


    # --- Metodi Gestione WebEngine, Processo, UI ---
    # [Metodi invariati inclusi di seguito]

    def _setup_webengine(self):
        """Inizializza la pagina WebEngine, il bridge e connette segnali."""
        page = WebEnginePage(self.view) if 'WebEnginePage' in globals() else QtWebEngineWidgets.QWebEnginePage(self.view)
        self.view.setPage(page)

        self.bridge = JSBridge() if 'JSBridge' in globals() else QtCore.QObject()
        self.channel = QtWebChannel.QWebChannel()
        if hasattr(self.bridge, 'receive'):
             logger.debug("Registering JSBridge object 'bridge' on web channel.")
             self.channel.registerObject('bridge', self.bridge)
        else:
             logger.warning("JSBridge object does not have 'receive' method, not registered.")

        current_page = self.view.page()
        if current_page:
            current_page.setWebChannel(self.channel)
            current_page.setBackgroundColor(QtCore.Qt.GlobalColor.transparent)
        else:
            logger.critical("Impossibile ottenere la pagina web dopo averla impostata.")
            QtWidgets.QMessageBox.critical(self, "Errore Critico", "Impossibile inizializzare componente Web.")


    @QtCore.pyqtSlot(bool)
    def _handle_page_load_finished(self, ok):
        """Gestisce il segnale loadFinished (SOLO per logica fetch)."""
        logger.debug(f"Page load finished signal received, success: {ok}")
        if self._current_fetch_context:
            context = self._current_fetch_context
            context['page_load_error'] = not ok
            if not ok:
                failed_url = self.view.url().toString()
                logger.error(f"Errore caricamento pagina '{failed_url}' (contesto fetch per codice '{context['code']}')")
                QtCore.QMetaObject.invokeMethod(
                    self, "_process_fetch_result", QtCore.Qt.ConnectionType.QueuedConnection,
                    QtCore.Q_ARG(str, context['code']),
                    QtCore.Q_ARG(str, 'Errore Caricamento Pagina'),
                    QtCore.Q_ARG(list, [])
                )


    @QtCore.pyqtSlot()
    @QtCore.pyqtSlot(QtCore.QUrl)
    @QtCore.pyqtSlot(bool)
    def _update_navigation_buttons_state(self, *args):
        """Aggiorna stato e colore icone dei pulsanti di navigazione web."""
        page = self.view.page()
        if page:
            history = page.history()
            can_go_back = history.canGoBack()
            self.btn_back.setEnabled(can_go_back)

            can_go_forward = history.canGoForward()
            self.btn_forward.setEnabled(can_go_forward)

            is_valid_url = not page.url().isEmpty() and page.url().isValid()
            self.btn_reload.setEnabled(is_valid_url)
        else:
            self.btn_back.setEnabled(False); self.btn_forward.setEnabled(False); self.btn_reload.setEnabled(False)

    def start_processing(self):
        """Avvia il processo: seleziona file, leggi codici, avvia worker thread."""
        logger.info("Avvio start_processing...")
        if self.thread is not None and self.thread.isRunning():
            logger.warning("Elaborazione già in corso."); return

        logger.debug("Apertura dialogo selezione file...")
        options = QtWidgets.QFileDialog.Option.DontUseNativeDialog if sys.platform == 'linux' else QtWidgets.QFileDialog.Option(0)
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Seleziona file Excel", "", "Excel Files (*.xlsx *.xls)", options=options )

        if not file_path:
            logger.info("Nessun file selezionato, operazione annullata.");
            self.file_label.setText("Nessun file selezionato")
            self.file_label.setToolTip("Seleziona un file per iniziare")
            return
        self.current_file_path = file_path; logger.info(f"File selezionato: {self.current_file_path}")
        self.file_label.setText(f"{os.path.basename(self.current_file_path)}")
        self.file_label.setToolTip(self.current_file_path)

        codes = []
        workbook = None
        logger.debug(f"Lettura codici da: {self.current_file_path}")
        try:
            workbook = load_workbook(filename=self.current_file_path, read_only=True, data_only=True, keep_vba=False)
            if not workbook.sheetnames: raise ValueError("Il file Excel non contiene fogli.")
            sheet = workbook.active
            logger.debug(f"Lettura foglio attivo: {sheet.title}")
            if sheet.max_row <= 1: raise ValueError("Il foglio Excel è vuoto o contiene solo l'intestazione.")

            header_row = sheet[1]
            ricerca_col_idx = -1
            for idx, cell in enumerate(header_row):
                if cell.value is not None and str(cell.value).strip().lower() == COL_RICERCA.lower():
                    ricerca_col_idx = idx + 1
                    break
            if ricerca_col_idx == -1: raise ValueError(f"Colonna '{COL_RICERCA}' non trovata.")
            logger.debug(f"Colonna '{COL_RICERCA}' trovata all'indice: {ricerca_col_idx}")

            code_count = 0
            for row_cells in sheet.iter_rows(min_row=2, min_col=ricerca_col_idx, max_col=ricerca_col_idx):
                cell = row_cells[0]
                if cell.value is not None:
                    code = str(cell.value).strip()
                    if code and code.lower() != 'nan':
                        codes.append(code); code_count += 1
            logger.info(f"Estrazione completata: trovati {code_count} codici validi.")

        except (InvalidFileException, FileNotFoundError, ValueError, Exception) as e:
            error_prefix = "Errore File Excel"; error_msg = ""
            if isinstance(e, InvalidFileException): error_msg = "File non valido o corrotto."; logger.error(error_msg, exc_info=True)
            elif isinstance(e, FileNotFoundError): error_msg = f"File non trovato: {self.current_file_path}"; logger.error(error_msg)
            elif isinstance(e, ValueError): error_msg = f"Errore contenuto/struttura: {e}"; logger.warning(error_msg)
            else: error_msg = "Errore imprevisto lettura Excel."; error_prefix="Errore Lettura"; logger.exception(error_msg)
            self._set_status_message(f"❌ {error_prefix}", False)
            QtWidgets.QMessageBox.critical(self, error_prefix, f"{error_msg}\nDettagli nel log.")
            self._reset_ui_after_processing()
            return
        finally:
            if workbook:
                try: workbook.close(); logger.debug("Workbook (read-only) chiuso.")
                except Exception as close_err: logger.warning(f"Errore chiusura workbook (read-only): {close_err}")

        if not codes:
            logger.warning("Nessun codice valido trovato nel file.")
            self._set_status_message("⚠️ Nessun codice trovato.", False)
            QtWidgets.QMessageBox.information(self, "Nessun Codice", f"Nessun codice valido trovato nella colonna '{COL_RICERCA}'.")
            self._reset_ui_after_processing()
            return

        # --- Avvio Thread Worker ---
        self.btn_start.setEnabled(False); self.btn_stop.setEnabled(True); self.btn_open.setEnabled(False)
        self._set_status_message("⏳ Avvio elaborazione...", True)
        if hasattr(self, 'log') and self.log: self.log.clear()
        self._reset_badges(); self.progress.setValue(0); self.progress_label.setText("0%")

        logger.debug(f"Avvio thread worker per {len(codes)} codici...")
        try:
            self.thread = QtCore.QThread(self); self.worker = Worker(codes)
            self.worker.moveToThread(self.thread)
            self.worker.progress.connect(self.update_progress); self.worker.statusUpdate.connect(self.update_status)
            self.worker.logUpdate.connect(self.update_log); self.worker.badgeUpdate.connect(self.update_badge_ui)
            self.worker.resultsReady.connect(self.handle_results)
            self.worker.requestFetch.connect(self.do_fetch_state_async, QtCore.Qt.ConnectionType.QueuedConnection)
            self.fetchCompleted.connect(self.worker.processFetchedResult)
            self.thread.started.connect(self.worker.run); self.worker.finished.connect(self.handle_thread_finished)
            self.worker.finished.connect(self.thread.quit); self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.thread.start(); logger.info("Thread worker avviato.")
        except Exception as e_thread:
            error_msg = "Errore creazione/avvio thread worker."
            logger.exception(error_msg); self._set_status_message(f"❌ Errore Thread", False)
            QtWidgets.QMessageBox.critical(self, "Errore Thread", f"{error_msg}\n{e_thread}\n\nConsultare il log.")
            self._reset_ui_after_processing()
        logger.debug("Fine start_processing.")


    def stop_processing(self):
        """Richiede l'interruzione del worker thread."""
        if self.worker and self.thread and self.thread.isRunning():
            logger.info("Richiesta interruzione elaborazione al worker...")
            self.worker.request_stop()
            self._set_status_message("⏳ Interruzione in corso...", False)
            self.btn_stop.setEnabled(False)
        else: logger.info("Nessuna elaborazione attiva da interrompere.")


    @QtCore.pyqtSlot(int, int)
    def update_progress(self, current_value, max_value):
        """Aggiorna la barra di progresso e l'etichetta %."""
        if max_value > 0:
            percentage = int((current_value / max_value) * 100)
            self.progress.setMaximum(max_value)
            self.progress.setValue(current_value)
            self.progress_label.setText(f"{percentage}%")
        else:
            self.progress.setMaximum(1); self.progress.setValue(0); self.progress_label.setText("0%")


    @QtCore.pyqtSlot(str)
    def update_status(self, message):
        """Aggiorna l'etichetta di stato."""
        self.status.setText(message)


    @QtCore.pyqtSlot(str)
    def update_log(self, message):
        """Aggiunge messaggio al widget di log e fa scroll."""
        if hasattr(self, 'log') and self.log:
            self.log.append(message)
            scrollbar = self.log.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())


    @QtCore.pyqtSlot(str, int)
    def update_badge_ui(self, badge_prefix, count):
        """Aggiorna contatore badge UI con animazioni."""
        if badge_prefix not in self._badge_widgets_map:
            logger.warning(f"Prefisso badge UI non trovato per '{badge_prefix}'")
            return
        card, label, opacity_effect = self._badge_widgets_map[badge_prefix]
        if not card or not label or not opacity_effect:
            logger.error(f"Widget/effetto badge mancante per {badge_prefix}")
            return

        new_text = f"{badge_prefix}: {count}"; old_text = label.text()
        was_visible = card.isVisible() and abs(opacity_effect.opacity() - 1.0) < 0.01

        if badge_prefix in self._active_badge_animations:
            try: self._active_badge_animations[badge_prefix].stop()
            except Exception: pass
            del self._active_badge_animations[badge_prefix]

        if old_text != new_text:
            label.setText(new_text)
            logger.debug(f"Aggiornato badge: {badge_prefix}, Conto: {count}")

        animation_group = None
        # CASO 1: Fade-in
        if count > 0 and not was_visible:
            if card not in self._badge_widgets: self.badge_layout.addWidget(card); self._badge_widgets.add(card)
            card.setVisible(True); opacity_effect.setOpacity(0.0)
            fade_in_anim = QtCore.QPropertyAnimation(opacity_effect, b"opacity", card)
            fade_in_anim.setDuration(350); fade_in_anim.setStartValue(0.0); fade_in_anim.setEndValue(1.0)
            fade_in_anim.setEasingCurve(QtCore.QEasingCurve.Type.InOutQuad)
            animation_group = fade_in_anim
        # CASO 2: Flash (senza pulse per ora)
        elif count > 0 and was_visible and old_text != new_text:
             card.setVisible(True); opacity_effect.setOpacity(1.0)
             color_anim = QtCore.QVariantAnimation(card); color_anim.setDuration(400)
             original_color = card.property("originalBgColor"); flash_color = QtGui.QColor(COLOR_LUMA_WHITE)
             def interpolate_color(start, end, progress):
                r = int(start.red() + (end.red() - start.red()) * progress); g = int(start.green() + (end.green() - start.green()) * progress); b = int(start.blue() + (end.blue() - start.blue()) * progress); return QtGui.QColor(r, g, b)
             def update_flash_color(value):
                 progress = value * 2 if value <= 0.5 else (1 - value) * 2
                 current_color = interpolate_color(original_color, flash_color, progress)
                 original_style = card.property("originalStyleSheet")
                 bg_color_str_original = f"background-color: {original_color.name()}"
                 bg_color_str_current = f"background-color: {current_color.name()}"
                 if original_style and bg_color_str_original in original_style:
                     new_style = original_style.replace(bg_color_str_original, bg_color_str_current)
                     card.setStyleSheet(new_style)
             color_anim.setStartValue(0.0); color_anim.setEndValue(1.0)
             color_anim.valueChanged.connect(update_flash_color)
             original_style_sheet = card.property("originalStyleSheet")
             if original_style_sheet: color_anim.finished.connect(lambda style=original_style_sheet: card.setStyleSheet(style))
             animation_group = color_anim
        # CASO 3: Fade-out
        elif count == 0 and was_visible:
            fade_out_anim = QtCore.QPropertyAnimation(opacity_effect, b"opacity", card)
            fade_out_anim.setDuration(350); fade_out_anim.setStartValue(1.0); fade_out_anim.setEndValue(0.0)
            fade_out_anim.finished.connect(lambda w=card: w.setVisible(False))
            animation_group = fade_out_anim

        if animation_group:
            self._active_badge_animations[badge_prefix] = animation_group
            animation_group.finished.connect(lambda bp=badge_prefix: self._active_badge_animations.pop(bp, None))
            animation_group.start(QtCore.QAbstractAnimation.DeletionPolicy.DeleteWhenStopped)
        elif count == 0 and not was_visible:
             card.setVisible(False); opacity_effect.setOpacity(0.0)


    @QtCore.pyqtSlot(list)
    def handle_results(self, results_list):
        """Riceve risultati e avvia salvataggio Excel."""
        logger.info(f"Ricevuti {len(results_list)} risultati finali dal worker.")
        if results_list and self.current_file_path:
            self._save_results_to_excel(results_list, self.current_file_path)
        elif not results_list and self.worker and not self.worker._stop_requested:
             logger.info("Nessun risultato valido raccolto."); self._set_status_message("ℹ️ Nessun risultato da salvare.", False)
        elif not self.current_file_path:
             logger.warning("Impossibile salvare: percorso file non definito."); self._set_status_message("⚠️ File non definito.", False)

        # Inserisci questo codice DENTRO la classe App in main_window.py,
        # SOSTITUENDO la funzione _save_results_to_excel esistente.

    def _save_results_to_excel(self, results_list, original_file_path):
            """
            Scrive i risultati nel file Excel. Se le colonne di output mancano,
            le crea automaticamente nella prima riga.
            Salva sull'originale se possibile, altrimenti crea una copia.
            """
            logger.info(
                f"Tentativo salvataggio {len(results_list)} risultati su {os.path.basename(original_file_path)}")
            self._set_status_message(f"⏳ Salvataggio risultati...", True)
            workbook = None
            output_file_path = original_file_path
            is_read_only_or_corrupted = False
            workbook_to_save = None
            sheet = None

            try:
                # 1. Carica il workbook (gestendo permessi/errori e creando copia se necessario)
                try:
                    workbook = openpyxl.load_workbook(filename=original_file_path)
                    logger.debug(f"Workbook '{os.path.basename(original_file_path)}' caricato per scrittura.")
                    is_read_only_or_corrupted = False
                    output_file_path = original_file_path
                except (PermissionError, IOError) as pe:
                    logger.warning(
                        f"Errore permesso/IO su '{os.path.basename(original_file_path)}': {pe}. Tento salvataggio su copia.")
                    is_read_only_or_corrupted = True
                    try:
                        # Ricarica in read-only per copiare i dati
                        workbook_ro = openpyxl.load_workbook(filename=original_file_path, read_only=True,
                                                             data_only=True)
                        if not workbook_ro.sheetnames: raise ValueError("File originale (read-only) non ha fogli.")
                        original_sheet = workbook_ro.active
                        workbook = openpyxl.Workbook()  # Crea nuovo workbook per la copia
                        sheet = workbook.active
                        sheet.title = original_sheet.title
                        # Copia intestazione e dati
                        for row_idx in range(1, original_sheet.max_row + 1):
                            row_values = [cell.value for cell in original_sheet[row_idx]]
                            sheet.append(row_values)
                        workbook_ro.close()  # Chiudi il read-only
                        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        base, ext = os.path.splitext(original_file_path)
                        output_file_path = f"{base}_output_{timestamp}{ext}"
                        logger.info(f"Salvataggio avverrà su nuovo file: {output_file_path}")
                        QtWidgets.QMessageBox.information(self, "Salvataggio su Copia",
                                                          f"Impossibile scrivere sul file originale.\n"
                                                          f"I risultati verranno salvati nel file:\n"
                                                          f"'{os.path.basename(output_file_path)}'")
                    except Exception as e_copy:
                        logger.error(f"Errore critico durante creazione copia file: {e_copy}", exc_info=True)
                        self._set_status_message("❌ Errore creazione copia.", False)
                        QtWidgets.QMessageBox.critical(self, "Errore Copia File",
                                                       f"Impossibile creare copia del file:\n{e_copy}")
                        return
                except (InvalidFileException, KeyError, Exception) as e_load:
                    # KeyError può accadere con file corrotti in openpyxl
                    logger.error(f"Errore caricamento workbook '{os.path.basename(original_file_path)}': {e_load}",
                                 exc_info=True)
                    is_read_only_or_corrupted = True  # Trattalo come non scrivibile
                    self._set_status_message("❌ Errore caricamento file.", False)
                    QtWidgets.QMessageBox.critical(self, "Errore File",
                                                   f"Impossibile caricare file Excel:\n'{os.path.basename(original_file_path)}'.\nPotrebbe essere corrotto o non valido.")
                    return

                # Assicurati di avere un foglio su cui lavorare
                if not sheet:  # Se non è stato creato dalla logica di copia
                    if not workbook or not workbook.sheetnames:
                        raise ValueError("Workbook non valido o senza fogli.")
                    sheet = workbook.active
                workbook_to_save = workbook  # Il workbook su cui effettivamente si salverà

                # --- NUOVA LOGICA: GESTIONE E MAPPATURA DIRETTA INTESTAZIONI ---
                header_row_idx = 1
                header_row = sheet[header_row_idx]
                # Leggi le intestazioni esistenti INIZIALMENTE
                existing_headers = {str(cell.value).strip().lower(): cell.column for cell in header_row if cell.value}

                # Mappa nomi config -> nomi excel effettivi (come prima)
                col_name_map = {
                    'Input Code': COL_RICERCA,
                    'Stato': COL_STATO,
                    'Protocollo uscita': COL_PROTOCOLLO,
                    'Provvedimento': COL_PROVVEDIMENTO,
                    'Data Provvedimento': COL_DATA_PROVV,
                    'Codice richiesta (risultato)': COL_CODICE_RIS,
                    'Note Usmaf': COL_NOTE
                }
                # Lista delle chiavi config per le colonne di output
                output_config_keys = [
                    'Stato', 'Protocollo uscita', 'Provvedimento',
                    'Data Provvedimento', 'Codice richiesta (risultato)', 'Note Usmaf'
                ]

                # Dizionario per memorizzare gli indici finali delle colonne
                final_col_indices = {}
                header_font = Font(bold=True)  # Font per nuove intestazioni

                # 1. Trova la colonna di ricerca (obbligatoria)
                ricerca_col_name_lower = COL_RICERCA.lower()
                if ricerca_col_name_lower in existing_headers:
                    final_col_indices['Input Code'] = existing_headers[ricerca_col_name_lower]
                    ricerca_col_idx = final_col_indices['Input Code']  # Memorizza per dopo
                else:
                    raise ValueError(f"Colonna di ricerca obbligatoria '{COL_RICERCA}' non trovata nella prima riga.")

                # Determina la prossima colonna libera *all'inizio*
                # Usiamo max_column qui, sperando sia affidabile. Se ci sono colonne nascoste potrebbe
                # ancora dare problemi, ma è lo standard openpyxl.
                next_available_col_idx = sheet.max_column + 1
                logger.debug(
                    f"Prossima colonna libera iniziale rilevata: {get_column_letter(next_available_col_idx)} ({next_available_col_idx})")

                # 2. Trova/Crea le colonne di output e mappa gli indici
                missing_added_log = []
                for config_key in output_config_keys:
                    excel_header_name = col_name_map[config_key]
                    excel_header_name_lower = excel_header_name.lower()

                    if excel_header_name_lower in existing_headers:
                        # Colonna già esistente, usa il suo indice
                        col_idx = existing_headers[excel_header_name_lower]
                        final_col_indices[config_key] = col_idx
                        logger.debug(
                            f"Intestazione '{excel_header_name}' trovata alla colonna {get_column_letter(col_idx)} ({col_idx})")
                    else:
                        # Colonna mancante: aggiungila alla prossima colonna libera
                        col_idx = next_available_col_idx
                        cell = sheet.cell(row=header_row_idx, column=col_idx, value=excel_header_name)
                        cell.font = header_font
                        final_col_indices[config_key] = col_idx
                        missing_added_log.append(f"'{excel_header_name}' in col {get_column_letter(col_idx)}")
                        logger.debug(
                            f"Intestazione '{excel_header_name}' AGGIUNTA alla colonna {get_column_letter(col_idx)} ({col_idx})")
                        # Incrementa l'indice per la prossima colonna da aggiungere
                        next_available_col_idx += 1

                if missing_added_log:
                    logger.info(f"Aggiunte intestazioni mancanti: {', '.join(missing_added_log)}.")

                # --- FINE NUOVA LOGICA INTESTAZIONI ---

                # Ora `final_col_indices` dovrebbe contenere la mappatura corretta
                # tra le chiavi di configurazione ('Stato', 'Protocollo', ecc.) e
                # l'indice numerico della colonna corrispondente nel foglio Excel.

                # Stampa log indici finali per verifica
                log_indices_str = ", ".join(
                    [f"'{k}': col {get_column_letter(v)}({v})" for k, v in final_col_indices.items()])
                logger.debug(f"Indici colonne finali determinati: {log_indices_str}")

                # Il codice successivo (da "# 3. Scrivi i dati...") rimane invariato
                # perché ora dovrebbe usare gli indici corretti da final_col_indices.

                # 3. Scrivi i dati nelle righe corrette usando gli indici finali
                results_map = {str(res['Input Code']).strip(): res for res in results_list}
                updated_rows = 0
                processed_keys = set()
                # Definisci stili comuni per le celle di output
                # MODIFICA QUI: wrap_text=False per testo lineare
                output_cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
                output_cell_number_format = numbers.FORMAT_TEXT

                for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                    cell_ricerca = sheet.cell(row=row_idx, column=ricerca_col_idx)
                    excel_code = str(cell_ricerca.value).strip() if cell_ricerca.value is not None else ''

                    if excel_code and excel_code in results_map:
                        result_data = results_map[excel_code]
                        processed_keys.add(excel_code)

                        # Scrivi i valori nelle colonne di output corrispondenti
                        for config_key, col_idx in final_col_indices.items():
                            # Salta la colonna di ricerca, aggiorniamo solo l'output
                            if config_key == 'Input Code':
                                continue

                            if config_key in result_data:
                                # Ottieni il valore originale
                                value_to_write = result_data.get(config_key, '')

                                # MODIFICA QUI: Imposta "NOTA USMAF" se la colonna è 'Note Usmaf'
                                # e il valore è vuoto (None, stringa vuota o solo spazi)
                                if config_key == 'Note Usmaf':
                                    # Controlla se il valore è considerato "vuoto"
                                    if value_to_write is None or not str(value_to_write).strip():
                                        value_to_write = "NOTA USMAF"
                                        logger.debug(
                                            f"Codice '{excel_code}': Nota Usmaf vuota, impostato valore di default.")

                                # Scrivi il valore (originale o modificato)
                                cell_to_write = sheet.cell(row=row_idx, column=col_idx)
                                cell_to_write.value = value_to_write
                                # Applica formattazione (con wrap_text=False)
                                cell_to_write.number_format = output_cell_number_format
                                cell_to_write.alignment = output_cell_alignment
                            else:
                                logger.warning(
                                    f"Chiave risultato '{config_key}' non trovata nei dati per codice '{excel_code}'. Cella non aggiornata.")

                        updated_rows += 1
                        # Rimuovi dal dizionario per efficienza (se ci sono duplicati nel file Excel)
                        # Se vuoi gestire i duplicati, decommenta la riga sotto:
                        # if excel_code in results_map: del results_map[excel_code]

                unmatched_codes = set(results_map.keys()) - processed_keys
                if unmatched_codes:
                    logger.warning(
                        f"{len(unmatched_codes)} codici dai risultati non trovati nella colonna '{COL_RICERCA}' del file Excel. Esempi: {list(unmatched_codes)[:5]}...")

                    # --- AGGIUNTA: Auto-Fit Larghezza Colonne ---
                logger.debug("Avvio auto-fit larghezza colonne...")
                for col_idx_numeric in range(1, sheet.max_column + 1):
                    column_letter = get_column_letter(col_idx_numeric)
                    max_length = 0
                    try:
                        # Itera su tutte le celle della colonna per trovare la lunghezza massima
                        for cell in sheet[column_letter]:
                            if cell.value:
                                # Considera la lunghezza della rappresentazione stringa del valore
                                # Aggiungi un piccolo margine se necessario (es. +2)
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length

                        # Imposta la larghezza della colonna (aggiungi un piccolo margine, es. + 2)
                        # Il valore della larghezza in openpyxl è approssimativo al numero di caratteri
                        adjusted_width = max_length + 2
                        sheet.column_dimensions[column_letter].width = adjusted_width
                        # logger.debug(f"Impostata larghezza colonna {column_letter} a {adjusted_width} (max_length={max_length})")

                    except Exception as e_width:
                        logger.warning(f"Errore durante l'impostazione larghezza colonna {column_letter}: {e_width}")
                logger.debug("Auto-fit larghezza colonne completato.")
                    # --- FINE AGGIUNTA ---

                # 4. Salva il workbook
                logger.info(f"Tentativo salvataggio modifiche su '{os.path.basename(output_file_path)}'...")
                try:
                    workbook_to_save.save(output_file_path)
                    logger.info(
                        f"Salvataggio '{os.path.basename(output_file_path)}' completato con successo. {updated_rows} righe aggiornate.")
                    self._set_status_message(f"✅ Risultati salvati.", False)
                    QtWidgets.QMessageBox.information(self, "Salvataggio Completato",
                                                      f"{updated_rows} righe aggiornate nel file:\n"
                                                      f"'{os.path.basename(output_file_path)}'")
                except Exception as e_save:
                    logger.error(f"Errore durante il salvataggio finale del file '{output_file_path}': {e_save}",
                                 exc_info=True)
                    self._set_status_message(f"❌ Errore salvataggio.", False)
                    QtWidgets.QMessageBox.critical(self, "Errore Salvataggio",
                                                   f"Impossibile salvare le modifiche nel file:\n"
                                                   f"'{os.path.basename(output_file_path)}'.\n"
                                                   f"Verificare che il file non sia aperto altrove e di avere i permessi.\n\n Dettaglio: {e_save}")

            except (InvalidFileException, FileNotFoundError, ValueError, RuntimeError, Exception) as e:
                error_msg = str(e) if isinstance(e, (InvalidFileException, FileNotFoundError, ValueError,
                                                     RuntimeError)) else "Errore imprevisto durante il processo di salvataggio Excel."
                logger.exception(f"Errore critico nel processo _save_results_to_excel: {error_msg}")
                self._set_status_message(f"❌ Errore salvataggio Excel", False)
                QtWidgets.QMessageBox.critical(self, "Errore Salvataggio Excel",
                                               f"Si è verificato un errore:\n{error_msg}\n\n"
                                               f"Controllare i log per maggiori dettagli.")
            finally:
                # 5. Chiudi il workbook in ogni caso
                if workbook_to_save:  # Prova a chiudere quello che stavamo per salvare
                    try:
                        workbook_to_save.close()
                        logger.debug("Workbook (workbook_to_save) chiuso.")
                    except Exception as ce:
                        logger.warning(f"Errore chiusura workbook (workbook_to_save): {ce}")
                elif workbook:  # Altrimenti chiudi quello originale caricato se esiste
                    try:
                        workbook.close()
                        logger.debug("Workbook (originale/copia) chiuso.")
                    except Exception as ce:
                        logger.warning(f"Errore chiusura workbook (originale/copia): {ce}")

                self.spinner.stopAnimation()  # Assicura che lo spinner si fermi


    @QtCore.pyqtSlot()
    def handle_thread_finished(self):
        """Pulisce stato dopo fine worker thread."""
        logger.info("Thread di elaborazione terminato.")
        final_status = self.status.text(); was_stopped = self.worker is not None and self.worker._stop_requested
        if "salvat" not in final_status.lower() and "errore" not in final_status.lower():
            final_status = "❌ Elaborazione Interrotta." if was_stopped else "ℹ️ Elaborazione terminata."
        self._set_status_message(final_status, False); self._reset_ui_after_processing()
        self.thread = None; self.worker = None; self._current_fetch_context = None
        logger.debug("Riferimenti Thread, Worker, Contesto Fetch puliti.")


    def _set_status_message(self, base_message, is_waiting):
        """Imposta status label e visibilità spinner."""
        self.status.setText(base_message)
        if is_waiting: self.spinner.startAnimation()
        else: self.spinner.stopAnimation()


    @QtCore.pyqtSlot(str)
    def do_fetch_state_async(self, code):
        """Avvia fetch per un codice (chiamato da worker)."""
        page = self.view.page()
        if not page: logger.critical(f"Errore: Pagina Web non disponibile (Codice: {code})."); self._process_fetch_result(code, 'Errore Pagina (Interno)', []); return
        if self.worker is None or self.worker._stop_requested: logger.info(f"Fetch {code} saltato - worker fermato/nullo."); return

        logger.info(f"Ricevuta richiesta fetch ASYNC per codice: {code}")
        self.last_cells = [''] * 11; self.last_cells[8] = str(code).strip()
        self._current_fetch_context = { 'code': code, 'attempt': 0, 'start_time': None, 'check_count': 0, 'null_check_count': 0, 'page_load_error': False }
        self._attempt_fetch(page, code)


    def _attempt_fetch(self, page, code):
        """Avvia un tentativo di fetch (gestisce retries)."""
        if not self._current_fetch_context or self._current_fetch_context['code'] != code: logger.debug(f"Contesto cambiato, skip fetch {code}"); return
        attempt = self._current_fetch_context['attempt']
        if self.worker is None or self.worker._stop_requested: logger.info(f"Fetch {code} annullato (Tent. {attempt+1}) - Interrotto."); self._process_fetch_result(code, 'Interrotto', self.last_cells); return
        if self._current_fetch_context['page_load_error']: logger.error(f"Fetch {code} fallito (Tent. {attempt+1}) - Errore Pagina."); self._process_fetch_result(code, 'Errore Caricamento Pagina', self.last_cells); return
        if attempt > self.max_retries: logger.warning(f"Timeout finale {code} dopo {attempt} tentativi."); self._process_fetch_result(code, 'Errore Timeout', self.last_cells); return

        if attempt > 0:
            logger.info(f"Attesa {DELAY_BETWEEN_RETRIES}ms prima di riprovare '{code}' ({attempt + 1}/{self.max_retries + 1})")
            self._set_status_message(f"⏳ Riprovo '{code}' ({attempt + 1})...", False)
            QtCore.QTimer.singleShot(DELAY_BETWEEN_RETRIES, lambda p=page, c=code: self._execute_js_input(p, c))
        else:
             logger.debug(f"Avvio tentativo 1 fetch {code}")
             self._execute_js_input(page, code)


    def _execute_js_input(self, page, code):
        """Esegue JS per inserire codice."""
        if not self._current_fetch_context or self._current_fetch_context['code'] != code: return
        attempt = self._current_fetch_context['attempt']
        if self.worker is None or self.worker._stop_requested or self._current_fetch_context['page_load_error']:
            status = 'Interrotto' if self.worker._stop_requested else 'Errore Caricamento Pagina'; self._process_fetch_result(code, status, self.last_cells); return

        try:
            logger.info(f"Inserimento codice {code} (tent. {attempt+1})...")
            self._set_status_message(f"Inserimento {code}...", True)
            escaped_code = str(code).strip().replace("'", "\\'").replace('"', '\\"')
            js_input = f"document.getElementById('codiceRichiesta').value='{escaped_code}';"
            page.runJavaScript(js_input, 0, lambda result, p=page, c=code: self._schedule_js_click(p, c))
        except Exception as js_error:
            logger.error(f"Errore eccezione runJavaScript (Input) per {code} ({attempt+1}): {js_error}", exc_info=True); self._set_status_message(f"❌ Errore JS Input", False)
            self._schedule_next_attempt(page, code)


    def _schedule_js_click(self, page, code):
        """Schedula esecuzione click JS."""
        if not self._current_fetch_context or self._current_fetch_context['code'] != code: return
        if not page or not hasattr(page, 'url'): logger.error(f"Errore: Pagina non valida schedula click {code}."); self._process_fetch_result(code, 'Errore Pagina (Interno)', self.last_cells); return
        QtCore.QTimer.singleShot(DELAY_AFTER_INPUT_JS, lambda: self._execute_js_click(page, code))


    def _execute_js_click(self, page, code):
        """Esegue JS per click."""
        if not self._current_fetch_context or self._current_fetch_context['code'] != code: return
        attempt = self._current_fetch_context['attempt']
        if not page or not hasattr(page, 'url'): logger.error(f"Errore: Pagina non valida esegui click {code}."); self._process_fetch_result(code, 'Errore Pagina (Interno)', self.last_cells); return
        if self.worker is None or self.worker._stop_requested or self._current_fetch_context['page_load_error']:
             status = 'Interrotto' if self.worker._stop_requested else 'Errore Caricamento Pagina'; self._process_fetch_result(code, status, self.last_cells); return

        try:
            logger.info(f"Click ricerca {code} (tent. {attempt+1})...")
            self._set_status_message(f"Click ricerca {code}...", True)
            js_click = "document.getElementById('cercaRichiestaNullaOstaBtn').click();"
            page.runJavaScript(js_click, 0, lambda result, p=page, c=code: self._schedule_first_check(p, c))
        except Exception as js_error:
            logger.error(f"Errore eccezione runJavaScript (Click) per {code} ({attempt+1}): {js_error}", exc_info=True); self._set_status_message(f"❌ Errore JS Click", False)
            self._schedule_next_attempt(page, code)


    def _schedule_first_check(self, page, code):
        """Schedula primo check risultati."""
        if not self._current_fetch_context or self._current_fetch_context['code'] != code: return
        if not page or not hasattr(page, 'url'): logger.error(f"Errore: Pagina non valida schedula check {code}."); self._process_fetch_result(code, 'Errore Pagina (Interno)', self.last_cells); return
        logger.info(f"Attesa risultati {code}..."); self._set_status_message(f"Attesa risultato {code}...", True)
        self._current_fetch_context['start_time'] = time.monotonic(); self._current_fetch_context['check_count'] = 0; self._current_fetch_context['null_check_count'] = 0
        QtCore.QTimer.singleShot(DELAY_AFTER_CLICK_JS, lambda: self._check_fetch_result(page, code))


    def _check_fetch_result(self, page, code):
        """Esegue JS per controllare risultati."""
        if not self._current_fetch_context or self._current_fetch_context['code'] != code: return
        if not page or not hasattr(page, 'url'): logger.error(f"Errore: Pagina non valida check risultati {code}."); self._process_fetch_result(code, 'Errore Pagina (Interno)', self.last_cells); return
        attempt = self._current_fetch_context['attempt']
        if self.worker is None or self.worker._stop_requested or self._current_fetch_context['page_load_error']:
             status = 'Interrotto' if self.worker._stop_requested else 'Errore Caricamento Pagina'; self._process_fetch_result(code, status, self.last_cells); return

        start_time = self._current_fetch_context.get('start_time')
        if start_time is None: logger.warning(f"start_time mancante {code}, reimposto."); start_time = time.monotonic(); self._current_fetch_context['start_time'] = start_time
        elapsed_time_ms = (time.monotonic() - start_time) * 1000
        if elapsed_time_ms > FETCH_TIMEOUT_MS: logger.warning(f"Timeout ({attempt + 1}) {code} durante check ({elapsed_time_ms:.0f} ms)"); self._schedule_next_attempt(page, code); return

        check_num = self._current_fetch_context['check_count'] + 1; status_msg = f"Controllo risultato {code} ({check_num})..."; logger.debug(status_msg)
        self._set_status_message(status_msg, True); logger.debug(f"Eseguo JS lettura celle (Tent. {attempt+1}, Check {check_num}) {code}")
        try: page.runJavaScript(ALL_CELLS_JS, 0, self._handle_js_evaluation_result)
        except Exception as eval_error:
             logger.error(f"Errore eccezione runJavaScript (Check) {code} (Tent. {attempt+1}): {eval_error}", exc_info=True); self._set_status_message(f"❌ Errore JS Check", False)
             self._schedule_next_attempt(page, code)


    def _handle_js_evaluation_result(self, result):
        """Gestisce risultato JS celle."""
        current_context = self._current_fetch_context;
        if not current_context: logger.warning(f"Risultato JS ricevuto senza contesto: {str(result)[:50]}..."); return
        code = current_context['code']; attempt = current_context['attempt']; start_time = current_context['start_time']
        page = self.view.page()
        if not page or not hasattr(page, 'url'): logger.error(f"Errore Pagina in handle JS result {code}."); self._process_fetch_result(code, 'Errore Pagina (Interno)', self.last_cells); return
        if self.worker is None or self.worker._stop_requested or current_context['page_load_error']:
            status = 'Interrotto' if self.worker._stop_requested else 'Errore Caricamento Pagina'; self._process_fetch_result(code, status, self.last_cells); return

        logger.debug(f"Ricevuto JS result {code} (Tent. {attempt+1}, Check {current_context['check_count']+1}): {str(result)[:100]}...")
        current_context['check_count'] += 1; cells = result; result_found = False; fetched_state = 'Errore Interno'; current_cells_data = []

        if cells is None:
            current_context['null_check_count'] += 1; logger.debug(f"JS result None {code} (Check {current_context['check_count']}, Nulls {current_context['null_check_count']})")
            if current_context['null_check_count'] >= MAX_NULL_CHECKS: logger.warning(f"Elemento non trovato (JS) {code}."); self._process_fetch_result(code, 'Elemento Non Trovato (JS)', self.last_cells); return
        elif isinstance(cells, list):
            current_context['null_check_count'] = 0; current_cells_data = [str(c).strip() if c else '' for c in cells]
            while len(current_cells_data) < 11: current_cells_data.append('')
            if len(current_cells_data) >= 1 and 'Nessun dato presente' in current_cells_data[0]:
                fetched_state = 'Non Trovato'; current_cells_data[2] = fetched_state; result_found = True; logger.info(f"'Nessun dato presente' {code}.")
            elif len(current_cells_data) > 8:
                code_in_table = current_cells_data[8]
                if code_in_table == str(code).strip(): fetched_state = current_cells_data[2] if current_cells_data[2] else 'Sconosciuto'; result_found = True; logger.info(f"Corrispondenza trovata {code}. Stato: '{fetched_state}'")
                else: logger.warning(f"Codice tabella ('{code_in_table}') != cercato ('{code}'). Riprovo check."); result_found = False
            else: logger.debug(f"Lista JS corta {code}. Riprovo check."); result_found = False
        else: logger.error(f"Risultato JS inatteso {code}: Tipo {type(cells)}"); self._process_fetch_result(code, 'Errore Risultato JS', self.last_cells); return

        if result_found: logger.info(f"Risultato finale acquisito {code} -> '{fetched_state}'"); self._process_fetch_result(code, fetched_state, current_cells_data); return

        if start_time is None: start_time = time.monotonic()
        elapsed_time_ms = (time.monotonic() - start_time) * 1000
        if elapsed_time_ms > FETCH_TIMEOUT_MS: logger.warning(f"Timeout {attempt+1} {code} dopo check JS fallito."); self._schedule_next_attempt(page, code); return

        logger.debug(f"Risultato {code} non pronto/valido, schedulo check..."); QtCore.QTimer.singleShot(FETCH_CHECK_INTERVAL_MS, lambda p=page, c=code: self._check_fetch_result(p, c))


    def _schedule_next_attempt(self, page, code):
         """Incrementa tentativo e schedula prossima chiamata fetch."""
         current_context = self._current_fetch_context;
         if not current_context or current_context['code'] != code: return
         if not page or not hasattr(page, 'url'): logger.error(f"Errore Pagina schedula prox tentativo {code}."); self._process_fetch_result(code, 'Errore Pagina (Interno)', self.last_cells); return
         next_attempt = current_context['attempt'] + 1; logger.debug(f"Schedulo tentativo {next_attempt} {code}")
         current_context['attempt'] = next_attempt; current_context['start_time'] = None; current_context['check_count'] = 0; current_context['null_check_count'] = 0; current_context['page_load_error'] = False
         QtCore.QTimer.singleShot(0, lambda p=page, c=code: self._attempt_fetch(p, c))


    def _process_fetch_result(self, code, state, cells):
        """Finalizza fetch: aggiorna UI, invia segnale a worker, pulisce contesto."""
        logger.info(f"Processando risultato finale '{code}': Stato='{state}'");
        if self._current_fetch_context and self._current_fetch_context['code'] == code:
            self._current_fetch_context = None; logger.debug(f"Contesto fetch '{code}' pulito.")
        else: logger.warning(f"Processo risultato '{code}', ma contesto attuale è '{self._current_fetch_context['code'] if self._current_fetch_context else 'None'}'.")

        self._set_status_message(f"Risultato {code}: {state}", False) # Aggiorna status UI
        if self.worker:
             logger.debug(f"Invio fetchCompleted '{code}' (Stato: {state}) a worker.")
             QtCore.QMetaObject.invokeMethod(self.worker, "processFetchedResult", QtCore.Qt.ConnectionType.QueuedConnection, QtCore.Q_ARG(str, code), QtCore.Q_ARG(str, state), QtCore.Q_ARG(list, cells))
        else: logger.warning(f"Worker non disponibile, risultato {code} non inviato.")


    def open_nsis_url_test(self):
        """Carica URL NSIS."""
        logger.info("Tentativo apertura URL NSIS...")
        try:
            page = self.view.page();
            if page:
                try: page.loadFinished.disconnect(self._handle_page_load_finished)
                except TypeError: pass
            else: logger.warning("Pagina non trovata prima di caricare URL.")
            self.load_real_url() # Carica direttamente URL reale
        except Exception as e: logger.exception("Eccezione in open_nsis_url_test"); QtWidgets.QMessageBox.critical(self, "Errore Apertura URL", f"Errore.\n{e}")


    def load_real_url(self):
        """Carica URL da config.py."""
        page = self.view.page();
        if not page: logger.error("Impossibile caricare URL: pagina Web non disponibile."); QtWidgets.QMessageBox.warning(self, "Errore Interno", "Componente Web non pronto."); return
        if self._current_fetch_context: self._current_fetch_context['page_load_error'] = False # Reset errore per nuovo load

        try:
            url_str = URL_NSIS; real_url = QtCore.QUrl(url_str); logger.info(f"Caricamento URL: {real_url.toString()}")
            if real_url.isValid():
                 try: page.loadFinished.disconnect(self._handle_page_load_finished)
                 except TypeError: pass
                 page.loadFinished.connect(self._handle_page_load_finished) # Riconnetti per fetch
                 page.load(real_url)
            else: logger.error(f"URL configurato non valido: {url_str}"); QtWidgets.QMessageBox.warning(self, "URL Non Valido", f"URL non valido:\n{url_str}")
        except Exception as e: logger.exception("Eccezione in load_real_url"); QtWidgets.QMessageBox.critical(self, "Errore Caricamento", f"Errore.\n{e}")


    def create_badge(self, icon, label_text, color_hex):
        """Crea badge con stile Luma."""
        frame = QtWidgets.QFrame(); text_color = COLOR_LUMA_GRAY_90
        try:
            bg_color = QtGui.QColor(color_hex)
            luminance = (0.299 * bg_color.red() + 0.587 * bg_color.green() + 0.114 * bg_color.blue()) / 255
            text_color = COLOR_LUMA_GRAY_90 if luminance > 0.5 else COLOR_LUMA_WHITE
        except Exception as e: logger.warning(f"Errore calcolo colore testo badge: {e}")

        frame.setStyleSheet(f""" QFrame {{ background-color: {color_hex}; border-radius: {LUMA_BORDER_RADIUS_BUTTON}; padding: 5px 10px; border: none; }} QLabel {{ color: {text_color}; font-weight: 500; font-size: 11px; background-color: transparent; border: none; padding: 0; margin: 0; font-family: "{self.ui_font_family}"; }} """)
        frame.setSizePolicy(QtWidgets.QSizePolicy.Policy.MinimumExpanding, QtWidgets.QSizePolicy.Policy.Fixed)
        layout = QtWidgets.QHBoxLayout(frame); layout.setContentsMargins(8, 4, 8, 4); layout.setSpacing(5)
        badge_prefix_text = f"{icon} {label_text}"
        label = QtWidgets.QLabel(f"{badge_prefix_text}: 0"); label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter); label.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred); layout.addWidget(label)
        opacity_effect = QtWidgets.QGraphicsOpacityEffect(frame); opacity_effect.setOpacity(0.0)
        frame.setGraphicsEffect(opacity_effect); frame.setProperty("opacityEffect", opacity_effect); frame.setProperty("badgePrefix", badge_prefix_text); frame.setProperty("originalStyleSheet", frame.styleSheet()); frame.setProperty("originalBgColor", QtGui.QColor(color_hex)); frame.setProperty("flashColor", QtGui.QColor(COLOR_LUMA_WHITE))
        return frame, label, opacity_effect


    def animate_badge(self, badge_frame):
         """Placeholder, logica ora in update_badge_ui."""
         pass

    def _reset_badges(self):
        """Nasconde badge, resetta contatori, stop animazioni."""
        logger.debug("Resetting badges UI...")
        for prefix, animation in list(self._active_badge_animations.items()):
            try: animation.stop()
            except Exception as e_stop: logger.warning(f"Errore stop animazione {prefix}: {e_stop}")
        self._active_badge_animations.clear()

        for widget in list(self._badge_widgets):
             if widget:
                 widget.setVisible(False)
                 opacity_effect = widget.property("opacityEffect"); original_style = widget.property("originalStyleSheet")
                 if isinstance(opacity_effect, QtWidgets.QGraphicsOpacityEffect): opacity_effect.setOpacity(0.0)
                 if original_style: widget.setStyleSheet(original_style)
                 self.badge_layout.removeWidget(widget)
        self._badge_widgets.clear()

        for prefix, (card, label, opacity_effect) in self._badge_widgets_map.items():
            if label: label.setText(f"{prefix}: 0")
            if card: card.setVisible(False);
            if isinstance(opacity_effect, QtWidgets.QGraphicsOpacityEffect): opacity_effect.setOpacity(0.0);
            original_style = card.property("originalStyleSheet")
            if original_style: card.setStyleSheet(original_style)
        logger.debug("Reset badges UI completato.")


    def _reset_ui_after_processing(self):
        """Resetta stato UI dopo elaborazione."""
        logger.debug("Resetting UI after processing...")
        self.btn_start.setEnabled(True); self.btn_stop.setEnabled(False); self.btn_open.setEnabled(True)
        self.spinner.stopAnimation()
        self.file_label.setText("Nessun file selezionato"); self.file_label.setToolTip("Seleziona un file per iniziare")


    def closeEvent(self, event: QtGui.QCloseEvent):
        """Gestisce evento chiusura finestra."""
        logger.info("Evento Chiusura Finestra Principale.")
        if self.thread is not None and self.thread.isRunning():
            reply = QtWidgets.QMessageBox.question(self, "Conferma Uscita", "Elaborazione in corso.\nInterrompere e uscire?", QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No, QtWidgets.QMessageBox.StandardButton.No)
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                logger.info("Interruzione richiesta prima della chiusura...")
                self.stop_processing()
                if self.thread and not self.thread.wait(500): logger.warning("Timeout attesa terminazione worker durante chiusura.")
                else: logger.info("Worker thread terminato (o timeout).")
                event.accept()
            else: logger.info("Chiusura annullata."); event.ignore()
        else: logger.info("Nessuna elaborazione in corso, chiusura normale."); event.accept()

# ------------------- Fine Classe App -------------------